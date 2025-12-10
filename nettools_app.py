#!/usr/bin/env python3
"""
NetTools Suite - IPv4 Scanner & MAC Formatter
Author: freakms
Company: ich schwöre feierlich ich bin ein tunichtgut
Version: 1.0.0

Modern desktop application for network utilities.
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import ipaddress
import csv
import re
import sys
import subprocess
import socket

# Import UI modules
from ui.dashboard_ui import DashboardUI
from ui.scanner_ui import ScannerUI
from ui.portscan_ui import PortScannerUI
from ui.dns_ui import DNSLookupUI
from ui.subnet_ui import SubnetCalculatorUI
from ui.mac_ui import MACFormatterUI
from ui.traceroute_ui import TracerouteUI
from ui.panos_ui import PANOSUI
from ui.bandwidth_ui import BandwidthUI
try:
    import telnetlib
    TELNETLIB_AVAILABLE = True
except ImportError:
    TELNETLIB_AVAILABLE = False
from datetime import datetime
from pathlib import Path
import platform
import os
from pythonping import ping
from PIL import Image, ImageDraw
import io
import json
import xml.etree.ElementTree as ET

# Try to import matplotlib (optional, used for Live Ping Monitor graphing)
try:
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    Figure = None
    FigureCanvasTkAgg = None

# Design system and UI components
from design_constants import COLORS, SPACING, RADIUS, FONTS, BUTTON_SIZES
from ui_components import (
    StyledCard, StyledButton, StyledEntry, SectionTitle, SubTitle,
    ResultRow, StatusBadge, SectionSeparator, LoadingSpinner, InfoBox, DataGrid
)

# Tool modules
from tools.scanner import IPv4Scanner
from tools.mac_formatter import OUILookup, MACFormatter
from tools.scan_manager import ScanManager
from tools.network_profile_manager import NetworkProfileManager
from tools.history_manager import HistoryManager
from tools.network_icon import NetworkIcon
from tools.live_ping_monitor import LivePingMonitor
from tools.bandwidth_tester import BandwidthTester
from tools.port_scanner import PortScanner
from tools.dns_lookup import DNSLookup
from tools.subnet_calculator import SubnetCalculator
from tools.traceroute import Traceroute
from tools.phpipam_tool import PHPIPAMTool

# phpIPAM integration
try:
    from phpipam_config import PHPIPAMConfig
    from phpipam_client import PHPIPAMClient
    PHPIPAM_AVAILABLE = True
except ImportError:
    PHPIPAM_AVAILABLE = False
    print("phpIPAM modules not found. Integration will be disabled.")

# Application metadata
APP_NAME = "NetTools Suite"
APP_VERSION = "1.11.0"
APP_COMPANY = "freakms - ich schwöre feierlich ich bin ein tunichtgut"

# Configure CustomTkinter
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")


# Classes moved to tools/ module for better organization
# - ScanManager -> tools/scan_manager.py
# - NetworkProfileManager -> tools/network_profile_manager.py
# - HistoryManager -> tools/history_manager.py
# - NetworkIcon -> tools/network_icon.py
# - IPv4Scanner -> tools/scanner.py
# - OUILookup, MACFormatter -> tools/mac_formatter.py


class NetToolsApp(ctk.CTk):
    """Main Application Window"""
    
    def __init__(self):
        super().__init__()
        
        # Window configuration
        self.title(f"{APP_NAME} - {APP_COMPANY}")
        self.geometry("1200x800")
        self.minsize(1000, 700)
        
        # Set icon
        try:
            icon_img = NetworkIcon.create_icon(256)
            # Save temporarily for tkinter
            icon_path = Path("temp_icon.png")
            icon_img.save(icon_path)
            self.iconphoto(True, ctk.CTkImage(light_image=icon_img, size=(32, 32))._light_image)
            if icon_path.exists():
                icon_path.unlink()
        except Exception:
            pass
        
        # Initialize scanner
        self.scanner = IPv4Scanner()
        self.scan_thread = None
        
        # Performance: Debounced updates
        self.update_buffer = []
        self.update_timer = None
        self.UPDATE_BATCH_SIZE = 10  # Update UI every 10 results
        self.UPDATE_INTERVAL_MS = 100  # Or every 100ms
        
        # Pagination for results
        self.all_results = []  # Store all results
        self.scan_current_page = 1
        self.results_per_page = 100
        self.scan_total_pages = 1
        
        # Window persistence - MUST be set before loading favorites
        self.config_file = Path.home() / '.nettools_config.json'
        
        # Favorites only (recent removed as not useful)
        self.favorite_tools = self.load_favorites()
        
        # Scan profiles
        self.scan_profiles = self.load_scan_profiles()
        
        # Initialize history manager
        self.history = HistoryManager()
        
        # Initialize scan manager
        self.scan_manager = ScanManager()
        
        # Initialize network profile manager
        self.profile_manager = NetworkProfileManager()
        
        # Load OUI database
        self.oui_database = self.load_oui_database()
        
        # Base window size for scaling calculations
        self.base_width = 1200
        self.base_height = 800
        self.current_scale = 1.0
        
        # Store font references for scaling
        self.base_font_size = 11
        self.title_font_size = 24
        self.subtitle_font_size = 12
        self.label_font_size = 12
        
        # Load window state
        self.load_window_state()
        
        # Create UI (order matters: status bar must be packed before main content)
        self.create_sidebar()
        self.create_status_bar()
        self.create_main_content()
        
        # Bind keyboard shortcuts
        self.bind('<Return>', self.on_enter_key)
        # Note: Ctrl+E for export is handled within scanner UI
        self.bind('<Control-k>', self.open_quick_switcher)  # Quick switcher
        
        # Bind window resize for auto-scaling
        self.bind('<Configure>', self.on_window_resize)
        
        # Save window state on close
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def load_window_state(self):
        """Load and apply saved window geometry"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    geometry = config.get('window_geometry')
                    if geometry:
                        self.geometry(geometry)
                        return
        except Exception as e:
            print(f"Could not load window state: {e}")
        
        # Default window size and position
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Set minimum size
        self.minsize(1200, 800)
        
        # Center window on screen
        x = (screen_width - self.base_width) // 2
        y = (screen_height - self.base_height) // 2
        self.geometry(f"{self.base_width}x{self.base_height}+{x}+{y}")
    
    def save_window_state(self):
        """Save current window geometry and preferences"""
        try:
            config = {}
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
            
            # Save geometry
            config['window_geometry'] = self.geometry()
            
            # Save favorites
            config['favorite_tools'] = list(self.favorite_tools)
            
            with open(self.config_file, 'w') as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Could not save window state: {e}")
    
    def load_favorites(self):
        """Load favorite tools from config"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    return set(config.get('favorite_tools', []))
        except Exception as e:
            print(f"Could not load favorites: {e}")
        return set()
    
    def load_scan_profiles(self):
        """Load saved scan profiles"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    return config.get('scan_profiles', {})
        except Exception as e:
            print(f"Could not load scan profiles: {e}")
        return {}
    
    def save_scan_profiles(self):
        """Save scan profiles to config"""
        try:
            config = {}
            if self.config_file.exists():
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
            
            config['scan_profiles'] = self.scan_profiles
            
            with open(self.config_file, 'w') as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Could not save scan profiles: {e}")
    
    def on_closing(self):
        """Handle window closing"""
        self.save_window_state()
        self.destroy()
    
    def open_quick_switcher(self, event=None):
        """Open quick tool switcher dialog (Ctrl+K)"""
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("Quick Tool Switcher")
        dialog.geometry("600x400")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 600) // 2
        y = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Content
        content = ctk.CTkFrame(dialog)
        content.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Search entry
        search_frame = ctk.CTkFrame(content, fg_color="transparent")
        search_frame.pack(fill="x", pady=(0, SPACING['md']))
        
        search_label = ctk.CTkLabel(
            search_frame,
            text="🔍 Search Tools:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        search_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        search_entry = StyledEntry(
            search_frame,
            placeholder_text="Type to search..."
        )
        search_entry.pack(fill="x")
        search_entry.focus()
        
        # Results scrollable frame
        results_frame = ctk.CTkScrollableFrame(content)
        results_frame.pack(fill="both", expand=True, pady=SPACING['md'])
        
        # All tools list
        all_tools = [
            ("scann", "scanning", "IPv4 Scanner", "📡 Scan network for devices"),
            ("port", "scanning", "Port Scanner", "🔌 Scan ports on devices"),
            ("trace", "scanning", "Traceroute", "🛣️ Trace network paths"),
            ("ping", "scanning", "Live Ping Monitor", "📊 Monitor ping in real-time"),
            ("bandwidth", "scanning", "Bandwidth Test", "⚡ Test network speed"),
            ("dns", "utilities", "DNS Lookup", "🌐 Lookup DNS records"),
            ("subnet", "utilities", "Subnet Calculator", "🧮 Calculate subnets"),
            ("mac", "utilities", "MAC Formatter", "💻 Format MAC addresses"),
            ("compare", "management", "Scan Comparison", "📊 Compare scan results"),
            ("profiles", "management", "Network Profiles", "📁 Manage network profiles"),
            ("panos", "automation", "PAN-OS Generator", "🔥 Generate firewall configs"),
            ("phpipam", "automation", "phpIPAM Integration", "📦 Integrate with phpIPAM"),
        ]
        
        tool_buttons = []
        
        def filter_tools(event=None):
            """Filter tools based on search"""
            search_text = search_entry.get().lower()
            
            # Clear results
            for widget in results_frame.winfo_children():
                widget.destroy()
            
            # Show matching tools
            for tool_id, category, name, description in all_tools:
                if (search_text in name.lower() or 
                    search_text in description.lower() or
                    search_text in tool_id):
                    
                    tool_btn = StyledButton(
                        results_frame,
                        text=f"{name}\n{description}",
                        command=lambda tid=tool_id: switch_and_close(tid),
                        size="large",
                        variant="neutral"
                    )
                    tool_btn.pack(fill="x", pady=SPACING['xs'])
                    tool_buttons.append((tool_btn, tool_id))
        
        def switch_and_close(tool_id):
            """Switch to tool and close dialog"""
            self.switch_tool(tool_id)
            dialog.destroy()
        
        def on_key(event):
            """Handle keyboard navigation"""
            if event.keysym == 'Escape':
                dialog.destroy()
            elif event.keysym == 'Down':
                # Focus next button
                pass
            elif event.keysym == 'Up':
                # Focus previous button
                pass
        
        # Bind events
        search_entry.bind('<KeyRelease>', filter_tools)
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        dialog.bind('<Key>', on_key)
        
        # Initial filter (show all)
        filter_tools()
        
        return "break"  # Prevent default handling
    
    def load_oui_database(self):
        """Load OUI database from JSON file"""
        try:
            # Handle both development and PyInstaller bundled environments
            if getattr(sys, 'frozen', False):
                # Running as compiled executable
                bundle_dir = Path(sys._MEIPASS)
            else:
                # Running as script
                bundle_dir = Path(__file__).parent
            
            oui_path = bundle_dir / "oui_database.json"
            
            if oui_path.exists():
                with open(oui_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                print(f"OUI database not found at: {oui_path}")
        except Exception as e:
            print(f"Could not load OUI database: {e}")
        return {}
    
    def lookup_vendor(self, mac_address):
        """Lookup vendor from MAC address OUI"""
        if not mac_address or len(mac_address) < 8:
            return "Unknown"
        
        # Extract OUI (first 3 octets) and normalize format
        # Handle different MAC formats
        cleaned_mac = mac_address.replace(':', '').replace('-', '').replace('.', '').upper()
        
        if len(cleaned_mac) >= 6:
            # Format as XX:XX:XX for lookup
            oui = f"{cleaned_mac[0:2]}:{cleaned_mac[2:4]}:{cleaned_mac[4:6]}"
            vendor = self.oui_database.get(oui, "Unknown Vendor")
            return vendor
        
        return "Unknown"
        
    def create_sidebar(self):
        """Create modern sidebar navigation with electric violet theme"""
        # Sidebar frame with dark violet background
        self.sidebar = ctk.CTkFrame(
            self, 
            width=250, 
            corner_radius=0,
            fg_color=COLORS['dashboard_card']
        )
        self.sidebar.pack(side="left", fill="y", padx=0, pady=0)
        self.sidebar.pack_propagate(False)
        
        # Logo/Title section
        logo_frame = ctk.CTkFrame(
            self.sidebar, 
            height=100, 
            corner_radius=0, 
            fg_color="transparent"
        )
        logo_frame.pack(fill="x", padx=0, pady=0)
        logo_frame.pack_propagate(False)
        
        title_label = ctk.CTkLabel(
            logo_frame,
            text="⚡ NetTools",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color=COLORS['electric_violet']
        )
        title_label.pack(padx=20, pady=(25, 5))
        
        subtitle_label = ctk.CTkLabel(
            logo_frame,
            text="Professional Suite",
            font=ctk.CTkFont(size=13),
            text_color=COLORS['neon_cyan']
        )
        subtitle_label.pack(padx=20, pady=(0, 10))
        
        # Separator with electric violet glow
        separator = ctk.CTkFrame(
            self.sidebar, 
            height=2, 
            corner_radius=0,
            fg_color=COLORS['electric_violet']
        )
        separator.pack(fill="x", padx=10, pady=10)
        
        # Scrollable navigation container
        nav_scroll = ctk.CTkScrollableFrame(self.sidebar, fg_color="transparent")
        nav_scroll.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Quick Access - Live Monitor (prominent button)
        self.live_monitor_btn = StyledButton(
            nav_scroll,
            text="📊 Live Ping Monitor",
            command=self.open_live_ping_monitor,
            size="large",
            variant="success"
        )
        self.live_monitor_btn.pack(fill="x", padx=10, pady=(5, 15))
        
        # Favorites section - create but don't pack yet (will pack when populated)
        self.favorites_frame = ctk.CTkFrame(nav_scroll, fg_color="transparent")
        self.favorites_label = ctk.CTkLabel(
            self.favorites_frame,
            text="⭐ FAVORITES",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLORS['neon_cyan'],
            anchor="w"
        )
        self.favorites_buttons_frame = ctk.CTkFrame(self.favorites_frame, fg_color="transparent")
        
        # Navigation organized by categories
        self.nav_buttons = {}
        
        # Category structure: (category_name, [(page_id, label, tooltip), ...])
        nav_categories = [
            ("🏠 HOME", [
                ("dashboard", "   Dashboard", "Overview and quick actions"),
            ]),
            ("🔍 NETWORK SCANNING", [
                ("scanner", "   IPv4 Scanner", "Scan network for active hosts"),
                ("portscan", "   Port Scanner", "Scan for open ports on hosts"),
                ("traceroute", "   Traceroute", "Trace network path to host"),
                ("bandwidth", "   Bandwidth Test", "Test network speed with iperf3"),
            ]),
            ("🛠 NETWORK TOOLS", [
                ("dns", "   DNS Lookup", "Resolve hostnames and IP addresses"),
                ("subnet", "   Subnet Calculator", "Calculate subnet information"),
                ("mac", "   MAC Formatter", "Format and analyze MAC addresses"),
            ]),
            ("📊 MANAGEMENT", [
                ("compare", "   Scan Comparison", "Compare network scan results"),
                ("profiles", "   Network Profiles", "Manage network profiles"),
            ]),
            ("🛡 ADVANCED", [
                ("panos", "   PAN-OS Generator", "Generate PAN-OS CLI commands"),
                ("phpipam", "   phpIPAM", "Manage IP addresses with phpIPAM"),
            ]),
        ]
        
        self.current_page = "dashboard"
        
        # Store first category for positioning
        self.first_category_label = None
        
        # Render navigation with categories
        for idx, (category_name, items) in enumerate(nav_categories):
            # Category header with electric violet theme
            category_label = ctk.CTkLabel(
                nav_scroll,
                text=category_name,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=COLORS['neon_cyan'],
                anchor="w"
            )
            category_label.pack(fill="x", padx=15, pady=(15, 5))
            
            # Store reference to first category
            if idx == 0:
                self.first_category_label = category_label
            
            # Category items with violet theme
            for page_id, label, tooltip in items:
                btn = ctk.CTkButton(
                    nav_scroll,
                    text=label,
                    command=lambda p=page_id: self.switch_tool(p),
                    height=40,
                    corner_radius=6,
                    anchor="w",
                    font=ctk.CTkFont(size=13, family="Segoe UI"),
                    fg_color="transparent",
                    text_color=("gray10", "gray90"),
                    hover_color=COLORS['dashboard_card_hover'],
                    border_width=0
                )
                btn.pack(fill="x", padx=12, pady=2)
                
                # Add right-click context menu for favorites
                btn.bind("<Button-3>", lambda e, tid=page_id: self.show_tool_context_menu(e, tid))
                
                self.nav_buttons[page_id] = btn
        
        # Update initial button state
        self.nav_buttons["scanner"].configure(fg_color=("gray75", "gray25"))
        
        # Add some bottom padding to scroll area
        bottom_padding = ctk.CTkFrame(nav_scroll, fg_color="transparent", height=20)
        bottom_padding.pack(fill="x")
        
        # Theme selector at bottom (fixed position outside scroll)
        theme_frame = ctk.CTkFrame(self.sidebar, corner_radius=0, fg_color="transparent")
        theme_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        
        theme_label = ctk.CTkLabel(
            theme_frame, 
            text="Theme", 
            font=ctk.CTkFont(size=12),
            text_color=COLORS['neon_cyan']
        )
        theme_label.pack(pady=(0, 5))
        
        self.theme_selector = ctk.CTkOptionMenu(
            theme_frame,
            values=["Dark", "Light"],
            command=self.change_theme,
            width=220,
            height=40,
            corner_radius=8,
            font=ctk.CTkFont(size=13),
            fg_color=COLORS['electric_violet'],
            button_color=COLORS['neon_cyan'],
            button_hover_color=COLORS['neon_cyan_hover']
        )
        self.theme_selector.pack()
        self.theme_selector.set("Dark")
        
        # Initialize favorites UI
        self.update_favorites_ui()
        self.update_nav_button_stars()
    
    def switch_page(self, page_id):
        """Switch between pages with lazy loading"""
        if page_id == self.current_page:
            return
        
        # Update button states with electric violet theme
        for btn_id, btn in self.nav_buttons.items():
            if btn_id == page_id:
                btn.configure(
                    fg_color=COLORS['dashboard_card_hover'],
                    border_width=2,
                    border_color=COLORS['electric_violet']
                )
            else:
                btn.configure(
                    fg_color="transparent",
                    border_width=0
                )
        
        # Hide all pages
        for page in self.pages.values():
            page.pack_forget()
        
        # Lazy load page content if not already loaded
        if page_id not in self.pages_loaded:
            if page_id not in self.pages:
                self.pages[page_id] = ctk.CTkFrame(self.main_content, corner_radius=0)
            
            # Load page content based on page_id
            if page_id == "dashboard":
                self.create_dashboard_content(self.pages[page_id])
            elif page_id == "scanner":
                self.create_scanner_content(self.pages[page_id])
            elif page_id == "mac":
                self.create_mac_content(self.pages[page_id])
            elif page_id == "compare":
                self.create_comparison_content(self.pages[page_id])
            elif page_id == "profiles":
                self.create_profiles_content(self.pages[page_id])
            elif page_id == "portscan":
                self.create_portscan_content(self.pages[page_id])
            elif page_id == "traceroute":
                TracerouteUI(self, self.pages[page_id])
            elif page_id == "dns":
                self.create_dns_content(self.pages[page_id])
            elif page_id == "panos":
                PANOSUI(self, self.pages[page_id])
            elif page_id == "subnet":
                self.create_subnet_content(self.pages[page_id])
            elif page_id == "phpipam":
                self.create_phpipam_content(self.pages[page_id])
            elif page_id == "bandwidth":
                BandwidthUI(self, self.pages[page_id])
            
            self.pages_loaded[page_id] = True
        
        # Show selected page
        self.pages[page_id].pack(fill="both", expand=True, padx=0, pady=0)
        self.current_page = page_id
        
        # Update status bar based on page
        if page_id == "dashboard":
            self.status_label.configure(text="Network Command Center - Ready")
        elif page_id == "scanner":
            self.status_label.configure(text="Ready to scan network")
        elif page_id == "mac":
            self.status_label.configure(text="Ready to format MAC address")
        elif page_id == "compare":
            self.status_label.configure(text="Compare network scans")
        elif page_id == "profiles":
            self.status_label.configure(text="Manage network interface profiles")
        elif page_id == "portscan":
            self.status_label.configure(text="Scan for open ports on target hosts")
        elif page_id == "traceroute":
            self.status_label.configure(text="Trace network path and analyze latency to target host")
        elif page_id == "dns":
            self.status_label.configure(text="Resolve hostnames and IP addresses")
        elif page_id == "panos":
            self.status_label.configure(text="Generate PAN-OS address object CLI commands")
        elif page_id == "subnet":
            self.status_label.configure(text="Calculate subnet information from CIDR")
        elif page_id == "phpipam":
            self.status_label.configure(text="Manage IP addresses with phpIPAM")
    
    
    def show_page(self, page_id):
        """Alias for switch_page() - used by dashboard"""
        self.switch_page(page_id)

    def create_main_content(self):
        """Create main content area with pages (lazy loading for faster startup)"""
        # Main content frame
        self.main_content = ctk.CTkFrame(self, corner_radius=0)
        self.main_content.pack(side="right", fill="both", expand=True, padx=0, pady=0)
        
        # Create pages dictionary (empty frames, content loaded on demand)
        self.pages = {}
        self.pages_loaded = {}  # Track which pages have been loaded
        
        # Pre-create the dashboard page for fast initial display
        self.pages["dashboard"] = ctk.CTkFrame(self.main_content, corner_radius=0)
        self.create_dashboard_content(self.pages["dashboard"])
        self.pages_loaded["dashboard"] = True
        
        # Show the initial page (dashboard)
        self.pages["dashboard"].pack(fill="both", expand=True, padx=0, pady=0)
    

    def create_dashboard_content(self, parent):
        """Create Dashboard home page with electric violet theme"""
        dashboard_ui = DashboardUI(self)
        dashboard_ui.create_content(parent)
    def create_scanner_content(self, parent):
        """Create IPv4 Scanner page content"""
        try:
            scanner_ui = ScannerUI(self)
            scanner_ui.create_content(parent)
        except Exception as e:
            # Show error if scanner fails to load
            error_label = ctk.CTkLabel(
                parent,
                text=f"Error loading scanner:\n{str(e)}",
                font=ctk.CTkFont(size=14),
                text_color="red"
            )
            error_label.pack(padx=20, pady=20)
            print(f"Scanner error: {e}")
            import traceback
            traceback.print_exc()
    def create_mac_content(self, parent):
        """Create MAC Formatter page content"""
        mac_ui = MACFormatterUI(self)
        mac_ui.create_content(parent)
    
    
    def create_comparison_content(self, parent):
        """Create Scan Comparison page content"""
        # Title
        title_label = ctk.CTkLabel(
            parent,
            text="Network Scan Comparison",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(padx=20, pady=(20, 10))
        
        # Description
        desc_label = ctk.CTkLabel(
            parent,
            text="Compare two network scans to see what devices have appeared, disappeared, or changed status.",
            font=ctk.CTkFont(size=12)
        )
        desc_label.pack(padx=20, pady=(0, 20))
        
        # Comparison button
        compare_btn = ctk.CTkButton(
            parent,
            text="Open Scan Comparison Tool",
            command=self.show_scan_comparison,
            width=250,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        compare_btn.pack(pady=20)
    
    def create_profiles_content(self, parent):
        """Create Network Profiles page content"""
        # Scrollable content area
        scrollable = ctk.CTkScrollableFrame(parent)
        scrollable.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Title
        title_label = ctk.CTkLabel(
            scrollable,
            text="Network Profile Manager",
            font=ctk.CTkFont(size=FONTS['title'], weight="bold")
        )
        title_label.pack(pady=(0, SPACING['xs']))
        
        # Subtitle
        subtitle_label = SubTitle(
            scrollable,
            text="Manage network interface configurations and quick-switch profiles"
        )
        subtitle_label.pack(pady=(0, SPACING['lg']))
        
        # Admin warning if not admin
        if not self.is_admin():
            warning_frame = InfoBox(
                scrollable,
                message="⚠️ Administrator privileges required to change network settings",
                box_type="warning"
            )
            warning_frame.pack(fill="x", pady=(0, SPACING['lg']))
        
        # Refresh button
        refresh_btn = StyledButton(
            scrollable,
            text="🔄 Refresh Interfaces",
            command=self.refresh_interfaces,
            size="large",
            variant="neutral"
        )
        refresh_btn.pack(pady=(0, SPACING['lg']))
        
        # Current Interfaces Section
        interfaces_title = SectionTitle(
            scrollable,
            text="Network Interfaces (Current Status)"
        )
        interfaces_title.pack(pady=(SPACING['md'], SPACING['md']), anchor="w")
        
        # Frame to hold interface cards
        self.interfaces_frame = ctk.CTkFrame(scrollable, fg_color="transparent")
        self.interfaces_frame.pack(fill="x", pady=(0, SPACING['lg']))
        
        # Separator
        separator = SectionSeparator(scrollable)
        separator.pack(fill="x", pady=SPACING['lg'])
        
        # Saved Profiles Section
        profiles_title = SectionTitle(
            scrollable,
            text="Saved Profiles"
        )
        profiles_title.pack(pady=(0, SPACING['md']), anchor="w")
        
        # New profile button
        new_profile_btn = StyledButton(
            scrollable,
            text="➕ Create New Profile",
            command=self.create_new_profile,
            size="large",
            variant="success"
        )
        new_profile_btn.pack(pady=(0, SPACING['lg']))
        
        # Frame to hold profile cards
        self.profiles_frame = ctk.CTkFrame(scrollable, fg_color="transparent")
        self.profiles_frame.pack(fill="x")
        
        # Load initial data
        self.refresh_interfaces()
        self.refresh_profiles()
    def create_portscan_content(self, parent):
        """Create Port Scanner page content"""
        try:
            portscan_ui = PortScannerUI(self)
            portscan_ui.create_content(parent)
        except Exception as e:
            error_label = ctk.CTkLabel(
                parent,
                text=f"Error loading port scanner:\n{str(e)}",
                font=ctk.CTkFont(size=14),
                text_color="red"
            )
            error_label.pack(padx=20, pady=20)
            print(f"Port Scanner error: {e}")
            import traceback
            traceback.print_exc()
    
    def create_dns_content(self, parent):
        """Create DNS Lookup page content"""
        dns_ui = DNSLookupUI(self)
        dns_ui.create_content(parent)
    
    def create_subnet_content(self, parent):
        """Create Subnet Calculator page content"""
        subnet_ui = SubnetCalculatorUI(self)
        subnet_ui.create_content(parent)
    
    
    def create_bandwidth_content(self, parent):
        """Create bandwidth testing page"""
        parent.pack(fill="both", expand=True)
        
        # Initialize bandwidth tester
        self.bandwidth_tester = BandwidthTester()
        
        # Check if iperf3 is available
        if not self.bandwidth_tester.is_iperf3_available():
            self.show_iperf3_not_installed(parent)
            return
        
        # Scroll container
        scroll = ctk.CTkScrollableFrame(parent)
        scroll.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Header
        header_card = StyledCard(scroll)
        header_card.pack(fill="x", pady=(0, SPACING['md']))
        
        title = ctk.CTkLabel(
            header_card,
            text="🚀 Bandwidth Testing (iperf3)",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold")
        )
        title.pack(padx=SPACING['lg'], pady=SPACING['lg'])
        
        subtitle = SubTitle(
            header_card,
            text="Test network throughput and speed using iperf3"
        )
        subtitle.pack(padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        # Test Configuration Card
        config_card = StyledCard(scroll)
        config_card.pack(fill="x", pady=(0, SPACING['md']))
        
        config_title = SectionTitle(config_card, text="Test Configuration")
        config_title.pack(anchor="w", padx=SPACING['lg'], pady=(SPACING['lg'], SPACING['xs']))
        
        # Server Host
        host_label = ctk.CTkLabel(config_card, text="iperf3 Server Host *", font=ctk.CTkFont(size=FONTS['body'], weight="bold"))
        host_label.pack(anchor="w", padx=SPACING['lg'], pady=(SPACING['md'], SPACING['xs']))
        
        self.bandwidth_host = StyledEntry(config_card, placeholder_text="e.g., iperf.example.com or 192.168.1.100")
        self.bandwidth_host.pack(fill="x", padx=SPACING['lg'], pady=(0, SPACING['md']))
        
        # Settings row
        settings_frame = ctk.CTkFrame(config_card, fg_color="transparent")
        settings_frame.pack(fill="x", padx=SPACING['lg'], pady=(0, SPACING['md']))
        
        # Port
        port_label = ctk.CTkLabel(settings_frame, text="Port:", font=ctk.CTkFont(size=FONTS['body']))
        port_label.pack(side="left", padx=(0, SPACING['xs']))
        
        self.bandwidth_port = StyledEntry(settings_frame, placeholder_text="5201", width=80)
        self.bandwidth_port.insert(0, "5201")
        self.bandwidth_port.pack(side="left", padx=(0, SPACING['md']))
        
        # Duration
        duration_label = ctk.CTkLabel(settings_frame, text="Duration (sec):", font=ctk.CTkFont(size=FONTS['body']))
        duration_label.pack(side="left", padx=(0, SPACING['xs']))
        
        self.bandwidth_duration = StyledEntry(settings_frame, placeholder_text="10", width=60)
        self.bandwidth_duration.insert(0, "10")
        self.bandwidth_duration.pack(side="left")
        
        # Test buttons
        btn_frame = ctk.CTkFrame(config_card, fg_color="transparent")
        btn_frame.pack(fill="x", padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        self.bandwidth_upload_btn = StyledButton(
            btn_frame,
            text="⬆ Upload Test",
            command=self.run_upload_test,
            size="large",
            variant="primary"
        )
        self.bandwidth_upload_btn.pack(side="left", fill="x", expand=True, padx=(0, SPACING['xs']))
        
        self.bandwidth_download_btn = StyledButton(
            btn_frame,
            text="⬇ Download Test",
            command=self.run_download_test,
            size="large",
            variant="success"
        )
        self.bandwidth_download_btn.pack(side="left", fill="x", expand=True, padx=(SPACING['xs'], 0))
        
        # Results Card
        results_card = StyledCard(scroll)
        results_card.pack(fill="both", expand=True)
        
        results_title = SectionTitle(results_card, text="Test Results")
        results_title.pack(anchor="w", padx=SPACING['lg'], pady=(SPACING['lg'], SPACING['xs']))
        
        self.bandwidth_results_frame = ctk.CTkFrame(results_card, fg_color="transparent")
        self.bandwidth_results_frame.pack(fill="both", expand=True, padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        # Empty state
        self.show_bandwidth_empty_state()
        
        # Info box
        info_card = StyledCard(scroll)
        info_card.pack(fill="x", pady=(SPACING['md'], 0))
        
        info_title = ctk.CTkLabel(
            info_card,
            text="ℹ️ About iperf3 Testing",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        info_title.pack(anchor="w", padx=SPACING['lg'], pady=(SPACING['md'], SPACING['xs']))
        
        info_text = ctk.CTkLabel(
            info_card,
            text="iperf3 requires a server to test against. You can:\n\n" +
                 "• Use a public iperf3 server (search online)\n" +
                 "• Set up your own server: iperf3 -s\n" +
                 "• Upload Test: Measures your upload speed to server\n" +
                 "• Download Test: Measures your download speed from server",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS['text_secondary'],
            justify="left"
        )
        info_text.pack(anchor="w", padx=SPACING['lg'], pady=(0, SPACING['md']))
    
    def show_iperf3_not_installed(self, parent):
        """Show iperf3 not installed message with instructions"""
        scroll = ctk.CTkScrollableFrame(parent)
        scroll.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Warning card
        warning_card = StyledCard(scroll)
        warning_card.pack(fill="both", expand=True)
        
        # Icon and title
        title = ctk.CTkLabel(
            warning_card,
            text="⚠️ iperf3 Not Installed",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold"),
            text_color=COLORS['warning']
        )
        title.pack(pady=(SPACING['lg'], SPACING['md']))
        
        # Message
        message = ctk.CTkLabel(
            warning_card,
            text="The bandwidth testing feature requires iperf3 to be installed on your system.",
            font=ctk.CTkFont(size=FONTS['body']),
            wraplength=600
        )
        message.pack(pady=(0, SPACING['lg']))
        
        # Installation instructions
        instructions_frame = ctk.CTkFrame(warning_card, fg_color=COLORS['bg_card'])
        instructions_frame.pack(fill="x", padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        inst_title = ctk.CTkLabel(
            instructions_frame,
            text="📥 Installation Instructions for Windows:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        inst_title.pack(anchor="w", padx=SPACING['md'], pady=(SPACING['md'], SPACING['xs']))
        
        instructions = """
Option 1: Using Chocolatey (Recommended)
  1. Install Chocolatey: https://chocolatey.org/install
  2. Open PowerShell as Administrator
  3. Run: choco install iperf3
  4. Restart this application

Option 2: Using Scoop
  1. Install Scoop: https://scoop.sh
  2. Open PowerShell
  3. Run: scoop install iperf3
  4. Restart this application

Option 3: Manual Installation
  1. Download from: https://iperf.fr/iperf-download.php
  2. Extract iperf3.exe to a folder
  3. Add folder to PATH environment variable
  4. Restart this application

Option 4: Using WSL (Windows Subsystem for Linux)
  1. Install WSL2
  2. In WSL terminal: sudo apt install iperf3
  3. Use from WSL terminal instead
        """
        
        inst_text = ctk.CTkLabel(
            instructions_frame,
            text=instructions,
            font=ctk.CTkFont(size=FONTS['small'], family="Consolas"),
            justify="left",
            anchor="w"
        )
        inst_text.pack(anchor="w", padx=SPACING['md'], pady=(0, SPACING['md']))
        
        # Refresh button
        refresh_btn = StyledButton(
            warning_card,
            text="🔄 Check Again",
            command=lambda: self.refresh_bandwidth_page(),
            size="large",
            variant="primary"
        )
        refresh_btn.pack(pady=(0, SPACING['lg']))
        
        # Alternative info
        alt_card = StyledCard(scroll)
        alt_card.pack(fill="x", pady=(SPACING['md'], 0))
        
        alt_title = ctk.CTkLabel(
            alt_card,
            text="💡 Alternative",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        alt_title.pack(anchor="w", padx=SPACING['lg'], pady=(SPACING['md'], SPACING['xs']))
        
        alt_text = ctk.CTkLabel(
            alt_card,
            text="You can also use online speed test websites like:\n" +
                 "• Speedtest.net\n" +
                 "• Fast.com\n" +
                 "• Google Speed Test (search 'internet speed test')",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS['text_secondary'],
            justify="left"
        )
        alt_text.pack(anchor="w", padx=SPACING['lg'], pady=(0, SPACING['md']))
    
    def refresh_bandwidth_page(self):
        """Refresh bandwidth page after iperf3 installation"""
        # Clear the page
        if "bandwidth" in self.pages:
            self.pages["bandwidth"].destroy()
            del self.pages["bandwidth"]
            self.pages_loaded.remove("bandwidth")
        
        # Reload the page
        self.switch_page("bandwidth")
    
    def show_bandwidth_empty_state(self):
        """Show empty state for bandwidth results"""
        for widget in self.bandwidth_results_frame.winfo_children():
            widget.destroy()
        
        empty_frame = ctk.CTkFrame(self.bandwidth_results_frame, fg_color="transparent")
        empty_frame.pack(fill="both", expand=True, pady=SPACING['xl'])
        
        empty_label = ctk.CTkLabel(
            empty_frame,
            text="📊\n\nNo test results yet\n\nRun an upload or download test to see results",
            font=ctk.CTkFont(size=FONTS['body']),
            text_color=COLORS['text_secondary'],
            justify="center"
        )
        empty_label.pack(expand=True)
    
    def run_upload_test(self):
        """Run upload speed test"""
        host = self.bandwidth_host.get().strip()
        port = self.bandwidth_port.get().strip() or "5201"
        duration = self.bandwidth_duration.get().strip() or "10"
        
        if not host:
            messagebox.showerror("Error", "Please enter an iperf3 server host")
            return
        
        try:
            port = int(port)
            duration = int(duration)
        except ValueError:
            messagebox.showerror("Error", "Port and duration must be numbers")
            return
        
        # Disable buttons during test
        self.bandwidth_upload_btn.configure(state="disabled", text="⬆ Testing...")
        self.bandwidth_download_btn.configure(state="disabled")
        
        self.show_bandwidth_testing()
        
        def callback(results, error):
            if error:
                messagebox.showerror("Test Failed", f"Upload test failed:\n{error}")
                self.show_bandwidth_empty_state()
            else:
                self.show_bandwidth_results(results, "Upload")
            
            self.bandwidth_upload_btn.configure(state="normal", text="⬆ Upload Test")
            self.bandwidth_download_btn.configure(state="normal")
        
        self.bandwidth_tester.test_client(host, port, duration, reverse=False, callback=callback)
    
    def run_download_test(self):
        """Run download speed test"""
        host = self.bandwidth_host.get().strip()
        port = self.bandwidth_port.get().strip() or "5201"
        duration = self.bandwidth_duration.get().strip() or "10"
        
        if not host:
            messagebox.showerror("Error", "Please enter an iperf3 server host")
            return
        
        try:
            port = int(port)
            duration = int(duration)
        except ValueError:
            messagebox.showerror("Error", "Port and duration must be numbers")
            return
        
        # Disable buttons during test
        self.bandwidth_upload_btn.configure(state="disabled")
        self.bandwidth_download_btn.configure(state="disabled", text="⬇ Testing...")
        
        self.show_bandwidth_testing()
        
        def callback(results, error):
            if error:
                messagebox.showerror("Test Failed", f"Download test failed:\n{error}")
                self.show_bandwidth_empty_state()
            else:
                self.show_bandwidth_results(results, "Download")
            
            self.bandwidth_upload_btn.configure(state="normal")
            self.bandwidth_download_btn.configure(state="normal", text="⬇ Download Test")
        
        self.bandwidth_tester.test_client(host, port, duration, reverse=True, callback=callback)
    
    def show_bandwidth_testing(self):
        """Show testing in progress"""
        for widget in self.bandwidth_results_frame.winfo_children():
            widget.destroy()
        
        testing_frame = ctk.CTkFrame(self.bandwidth_results_frame, fg_color="transparent")
        testing_frame.pack(fill="both", expand=True, pady=SPACING['xl'])
        
        testing_label = ctk.CTkLabel(
            testing_frame,
            text="⏳\n\nTest in progress...\n\nPlease wait",
            font=ctk.CTkFont(size=FONTS['body']),
            justify="center"
        )
        testing_label.pack(expand=True)
    
    def show_bandwidth_results(self, results, test_type):
        """Display bandwidth test results"""
        for widget in self.bandwidth_results_frame.winfo_children():
            widget.destroy()
        
        summary = self.bandwidth_tester.get_summary(results)
        
        if not summary:
            self.show_bandwidth_empty_state()
            return
        
        # Test type header
        type_label = ctk.CTkLabel(
            self.bandwidth_results_frame,
            text=f"✅ {test_type} Test Complete",
            font=ctk.CTkFont(size=FONTS['subheading'], weight="bold"),
            text_color=COLORS['success']
        )
        type_label.pack(pady=(0, SPACING['md']))
        
        # Main speed display
        if test_type == "Upload":
            speed = summary['sent_mbps']
        else:
            speed = summary['received_mbps']
        
        speed_card = ctk.CTkFrame(self.bandwidth_results_frame, fg_color=COLORS['bg_card'])
        speed_card.pack(fill="x", pady=(0, SPACING['md']))
        
        speed_label = ctk.CTkLabel(
            speed_card,
            text=f"{speed:.2f} Mbps",
            font=ctk.CTkFont(size=36, weight="bold")
        )
        speed_label.pack(pady=SPACING['lg'])
        
        # Detailed results
        details_data = [
            ("Sent (Upload)", f"{summary['sent_mbps']:.2f} Mbps", f"{summary['sent_bytes'] / 1_000_000:.2f} MB"),
            ("Received (Download)", f"{summary['received_mbps']:.2f} Mbps", f"{summary['received_bytes'] / 1_000_000:.2f} MB"),
            ("Local CPU Usage", f"{summary['cpu_utilization_local']:.1f}%", ""),
            ("Remote CPU Usage", f"{summary['cpu_utilization_remote']:.1f}%", ""),
        ]
        
        for label, value1, value2 in details_data:
            row = ctk.CTkFrame(self.bandwidth_results_frame, fg_color="transparent")
            row.pack(fill="x", pady=SPACING['xs'])
            
            label_widget = ctk.CTkLabel(row, text=f"{label}:", font=ctk.CTkFont(size=FONTS['body'], weight="bold"), width=150, anchor="w")
            label_widget.pack(side="left")
            
            value_widget = ctk.CTkLabel(row, text=value1, font=ctk.CTkFont(size=FONTS['body']), anchor="w")
            value_widget.pack(side="left", padx=SPACING['sm'])
            
            if value2:
                value2_widget = ctk.CTkLabel(row, text=value2, font=ctk.CTkFont(size=FONTS['small']), text_color=COLORS['text_secondary'], anchor="w")
                value2_widget.pack(side="left")
    
    def create_status_bar(self):
        """Create status bar with electric violet theme"""
        status_frame = ctk.CTkFrame(
            self, 
            height=35, 
            corner_radius=0,
            fg_color=COLORS['dashboard_card']
        )
        status_frame.pack(fill="x", side="bottom")
        status_frame.pack_propagate(False)
        
        self.status_label = ctk.CTkLabel(
            status_frame,
            text="Ready.",
            font=ctk.CTkFont(size=11),
            text_color=COLORS['neon_cyan']
        )
        self.status_label.pack(side="left", padx=15, pady=5)
        
        self.progress_bar = ctk.CTkProgressBar(
            status_frame, 
            width=200,
            progress_color=COLORS['electric_violet']
        )
        self.progress_bar.pack(side="left", padx=15, pady=5)
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()  # Hide initially
        
        copyright_label = ctk.CTkLabel(
            status_frame,
            text=f"© {APP_COMPANY}",
            font=ctk.CTkFont(size=10),
            text_color=COLORS['text_secondary']
        )
        copyright_label.pack(side="right", padx=15, pady=5)
    
    def is_admin(self):
        """Check if running with administrator privileges"""
        try:
            if platform.system() == "Windows":
                import ctypes
                return ctypes.windll.shell32.IsUserAnAdmin() != 0
            return os.geteuid() == 0
        except:
            return False
    
    def run_subprocess(self, cmd, **kwargs):
        """Run subprocess without showing console window (fixes ghost CMD windows)"""
        if platform.system() == "Windows":
            # Add flag to hide console window on Windows
            if 'creationflags' not in kwargs:
                kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW
        return subprocess.run(cmd, **kwargs)
    
    def restart_as_admin(self):
        """Restart application with administrator privileges (Windows UAC)"""
        try:
            if platform.system() == "Windows":
                import ctypes
                
                # Get the path to the executable or script
                if getattr(sys, 'frozen', False):
                    # Running as exe
                    script = sys.executable
                else:
                    # Running as script
                    script = os.path.abspath(sys.argv[0])
                
                # Trigger UAC elevation
                ctypes.windll.shell32.ShellExecuteW(
                    None,
                    "runas",
                    sys.executable if getattr(sys, 'frozen', False) else "python",
                    script if not getattr(sys, 'frozen', False) else "",
                    None,
                    1  # SW_SHOWNORMAL
                )
                
                # Close current instance
                self.quit()
        except Exception as e:
            messagebox.showerror("Error", f"Could not restart with admin privileges:\n{str(e)}")
    
    def get_network_interfaces(self):
        """Get list of network interfaces (Windows)"""
        try:
            result = self.run_subprocess(
                ["netsh", "interface", "ipv4", "show", "interfaces"],
                capture_output=True,
                text=True,
                timeout=5
            )
            
            interfaces = []
            lines = result.stdout.split('\n')[3:]  # Skip header lines
            
            for line in lines:
                if line.strip():
                    parts = line.split()
                    if len(parts) >= 4:
                        idx = parts[0]
                        status = parts[2]
                        name = ' '.join(parts[3:])
                        interfaces.append({
                            "index": idx,
                            "name": name,
                            "status": status
                        })
            
            return interfaces
        except Exception as e:
            print(f"Error getting interfaces: {e}")
            return []
    
    def get_interface_config(self, interface_name):
        """Get current IP configuration for an interface"""
        try:
            result = self.run_subprocess(
                ["netsh", "interface", "ipv4", "show", "config", interface_name],
                capture_output=True,
                text=True,
                timeout=5
            )
            
            config = {
                "dhcp": False,
                "ip": None,
                "subnet": None,
                "gateway": None,
                "dns": []
            }
            
            lines = result.stdout.split('\n')
            for i, line in enumerate(lines):
                if "DHCP enabled" in line and "Yes" in line:
                    config["dhcp"] = True
                elif "IP Address:" in line and i + 1 < len(lines):
                    config["ip"] = lines[i + 1].strip()
                elif "Subnet Prefix:" in line and i + 1 < len(lines):
                    subnet_line = lines[i + 1].strip()
                    config["subnet"] = subnet_line.split('/')[0] if '/' in subnet_line else subnet_line
                elif "Default Gateway:" in line and i + 1 < len(lines):
                    config["gateway"] = lines[i + 1].strip()
                elif "DNS servers configured through DHCP:" in line:
                    j = i + 1
                    while j < len(lines) and lines[j].strip() and not lines[j].startswith('Configuration'):
                        dns = lines[j].strip()
                        if dns:
                            config["dns"].append(dns)
                        j += 1
                elif "Statically Configured DNS Servers:" in line:
                    j = i + 1
                    while j < len(lines) and lines[j].strip() and not lines[j].startswith('Configuration'):
                        dns = lines[j].strip()
                        if dns:
                            config["dns"].append(dns)
                        j += 1
            
            return config
        except Exception as e:
            print(f"Error getting interface config: {e}")
            return None
    
    def refresh_interfaces(self):
        """Refresh the interface list display"""
        # Clear existing
        for widget in self.interfaces_frame.winfo_children():
            widget.destroy()
        
        interfaces = self.get_network_interfaces()
        
        if not interfaces:
            no_interfaces_label = ctk.CTkLabel(
                self.interfaces_frame,
                text="No network interfaces found",
                font=ctk.CTkFont(size=12)
            )
            no_interfaces_label.pack(pady=20)
            return
        
        for interface in interfaces:
            self.create_interface_card(interface)
    
    def create_interface_card(self, interface):
        """Create a card for an interface"""
        card = ctk.CTkFrame(self.interfaces_frame, corner_radius=8)
        card.pack(fill="x", pady=5)
        
        # Header with name and status
        header_frame = ctk.CTkFrame(card, fg_color="transparent")
        header_frame.pack(fill="x", padx=15, pady=(10, 5))
        
        name_label = ctk.CTkLabel(
            header_frame,
            text=interface["name"],
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="w"
        )
        name_label.pack(side="left")
        
        status_color = "#4CAF50" if interface["status"] == "connected" else "#757575"
        status_label = ctk.CTkLabel(
            header_frame,
            text=interface["status"].upper(),
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=status_color
        )
        status_label.pack(side="right")
        
        # Get current config
        config = self.get_interface_config(interface["name"])
        
        if config:
            # IP info
            if config["dhcp"]:
                ip_text = "DHCP Enabled"
                if config["ip"]:
                    ip_text += f" (Current IP: {config['ip']})"
            else:
                ip_text = f"Static IP: {config['ip'] or 'Not configured'}"
            
            ip_label = ctk.CTkLabel(
                card,
                text=ip_text,
                font=ctk.CTkFont(size=12),
                anchor="w"
            )
            ip_label.pack(fill="x", padx=15, pady=(0, 10))
            
            # Action buttons
            button_frame = ctk.CTkFrame(card, fg_color="transparent")
            button_frame.pack(fill="x", padx=15, pady=(0, 10))
            
            if config["dhcp"]:
                static_btn = ctk.CTkButton(
                    button_frame,
                    text="Set Static IP",
                    command=lambda i=interface["name"]: self.show_static_ip_dialog(i),
                    width=140,
                    height=36
                )
                static_btn.pack(side="left", padx=(0, 5))
            else:
                dhcp_btn = ctk.CTkButton(
                    button_frame,
                    text="Switch to DHCP",
                    command=lambda i=interface["name"]: self.set_dhcp(i),
                    width=140,
                    height=36,
                    fg_color=("#4CAF50", "#388E3C")
                )
                dhcp_btn.pack(side="left", padx=(0, 5))
                
                edit_btn = ctk.CTkButton(
                    button_frame,
                    text="Edit Static IP",
                    command=lambda i=interface["name"]: self.show_static_ip_dialog(i),
                    width=140,
                    height=36
                )
                edit_btn.pack(side="left", padx=(0, 5))
    
    def change_theme(self, theme):
        """Change application theme"""
        ctk.set_appearance_mode(theme)
    
    def _flush_update_buffer(self):
        """Flush buffered updates to UI"""
        if self.update_timer:
            self.after_cancel(self.update_timer)
            self.update_timer = None
        
        if not self.update_buffer:
            return
        
        # Process all buffered updates
        for completed, total, result in self.update_buffer:
            self._update_scan_progress(completed, total, result)
        
        self.update_buffer.clear()
    
    def show_export_dialog(self):
        """Show export options dialog"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Export Options")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 500) // 2
        y = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Content
        content = ctk.CTkFrame(dialog)
        content.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Title
        title = ctk.CTkLabel(
            content,
            text="📤 Export Scan Results",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold")
        )
        title.pack(pady=(0, SPACING['lg']))
        
        # Export scope
        scope_frame = ctk.CTkFrame(content, fg_color="transparent")
        scope_frame.pack(fill="x", pady=SPACING['md'])
        
        scope_label = ctk.CTkLabel(
            scope_frame,
            text="Export Scope:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        scope_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        scope_var = ctk.StringVar(value="all")
        
        all_radio = ctk.CTkRadioButton(
            scope_frame,
            text=f"All Results ({len(self.all_results)} total)",
            variable=scope_var,
            value="all"
        )
        all_radio.pack(anchor="w", pady=SPACING['xs'])
        
        current_page_radio = ctk.CTkRadioButton(
            scope_frame,
            text=f"Current Page Only ({len(self.result_rows)} results)",
            variable=scope_var,
            value="page"
        )
        current_page_radio.pack(anchor="w", pady=SPACING['xs'])
        
        online_radio = ctk.CTkRadioButton(
            scope_frame,
            text=f"Online Hosts Only ({sum(1 for r in self.all_results if r.get('status') == 'Online')} results)",
            variable=scope_var,
            value="online"
        )
        online_radio.pack(anchor="w", pady=SPACING['xs'])
        
        # Format selection
        format_frame = ctk.CTkFrame(content, fg_color="transparent")
        format_frame.pack(fill="x", pady=SPACING['md'])
        
        format_label = ctk.CTkLabel(
            format_frame,
            text="Export Format:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        format_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        format_var = ctk.StringVar(value="csv")
        
        formats = [
            ("csv", "CSV - Comma Separated Values"),
            ("json", "JSON - JavaScript Object Notation"),
            ("xlsx", "Excel - Microsoft Excel (requires openpyxl)"),
            ("html", "HTML - Web Page Report"),
            ("txt", "TXT - Plain Text"),
            ("xml", "XML - Extensible Markup Language"),
        ]
        
        for value, label in formats:
            radio = ctk.CTkRadioButton(
                format_frame,
                text=label,
                variable=format_var,
                value=value
            )
            radio.pack(anchor="w", pady=SPACING['xs'])
        
        # Buttons
        button_frame = ctk.CTkFrame(content, fg_color="transparent")
        button_frame.pack(fill="x", pady=SPACING['lg'], side="bottom")
        
        cancel_btn = StyledButton(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            size="medium",
            variant="neutral"
        )
        cancel_btn.pack(side="right", padx=SPACING['xs'])
        
        def do_export():
            scope = scope_var.get()
            format_type = format_var.get()
            dialog.destroy()
            self.perform_export(scope, format_type)
        
        export_btn = StyledButton(
            button_frame,
            text="📤 Export",
            command=do_export,
            size="large",
            variant="primary"
        )
        export_btn.pack(side="right")
    
    def perform_export(self, scope, format_type):
        """Perform the actual export"""
        # Filter results based on scope
        if scope == "all":
            results_to_export = self.all_results
        elif scope == "page":
            results_to_export = [row.result_data for row in self.result_rows if hasattr(row, 'result_data')]
        elif scope == "online":
            results_to_export = [r for r in self.all_results if r.get('status') == 'Online']
        else:
            results_to_export = self.all_results
        
        if not results_to_export:
            messagebox.showinfo("Information", "No data to export with selected filter.")
            return
        
        # Get desktop path
        desktop = Path.home() / "Desktop"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"NetToolsScan_{timestamp}"
        
        # File extension mapping
        ext_map = {
            "csv": ".csv",
            "json": ".json",
            "xlsx": ".xlsx",
            "html": ".html",
            "txt": ".txt",
            "xml": ".xml"
        }
        
        # Ask for save location
        filepath = filedialog.asksaveasfilename(
            defaultextension=ext_map.get(format_type, ".csv"),
            filetypes=[(f"{format_type.upper()} files", f"*{ext_map.get(format_type, '.csv')}")],
            initialdir=desktop,
            initialfile=default_filename
        )
        
        if not filepath:
            return
        
        try:
            if format_type == "csv":
                self._export_as_csv(filepath, results_to_export)
            elif format_type == "json":
                self._export_as_json(filepath, results_to_export)
            elif format_type == "xlsx":
                self._export_as_excel(filepath, results_to_export)
            elif format_type == "html":
                self._export_as_html(filepath, results_to_export)
            elif format_type == "txt":
                self._export_as_txt(filepath, results_to_export)
            elif format_type == "xml":
                self._export_as_xml(filepath, results_to_export)
            
            messagebox.showinfo(
                "Export Successful",
                f"{len(results_to_export)} results exported to:\n{filepath}"
            )
        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Error exporting scan results: {str(e)}"
            )
    
    def _export_as_csv(self, filepath, results):
        """Export results as CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['IP Address', 'Hostname', 'Status', 'Response Time'])
            for result in results:
                writer.writerow([
                    result.get('ip', ''),
                    result.get('hostname', ''),
                    result.get('status', ''),
                    result.get('rtt', '')
                ])
    
    def _export_as_json(self, filepath, results):
        """Export results as JSON"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
    
    def _export_as_excel(self, filepath, results):
        """Export results as Excel (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scan Results"
            
            # Header
            headers = ['IP Address', 'Hostname', 'Status', 'Response Time']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for result in results:
                ws.append([
                    result.get('ip', ''),
                    result.get('hostname', ''),
                    result.get('status', ''),
                    result.get('rtt', '')
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2
            
            wb.save(filepath)
        except ImportError:
            messagebox.showerror(
                "Module Not Found",
                "Excel export requires 'openpyxl' module.\n\nInstall with: pip install openpyxl"
            )
    
    def _export_as_html(self, filepath, results):
        """Export results as HTML report"""
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>NetTools Scan Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        h1 {{ color: #333; }}
        .info {{ background: #e3f2fd; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; background: white; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        th {{ background: #4472C4; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        tr:hover {{ background: #f5f5f5; }}
        .online {{ color: green; font-weight: bold; }}
        .offline {{ color: red; }}
    </style>
</head>
<body>
    <h1>🔍 NetTools Scan Report</h1>
    <div class="info">
        <strong>Generated:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
        <strong>Total Results:</strong> {len(results)}<br>
        <strong>Online Hosts:</strong> {sum(1 for r in results if r.get('status') == 'Online')}<br>
        <strong>Offline Hosts:</strong> {sum(1 for r in results if r.get('status') == 'Offline')}
    </div>
    <table>
        <thead>
            <tr>
                <th>IP Address</th>
                <th>Hostname</th>
                <th>Status</th>
                <th>Response Time</th>
            </tr>
        </thead>
        <tbody>
"""
        for result in results:
            status_class = "online" if result.get('status') == 'Online' else "offline"
            html_content += f"""
            <tr>
                <td>{result.get('ip', '')}</td>
                <td>{result.get('hostname', '-')}</td>
                <td class="{status_class}">{result.get('status', '')}</td>
                <td>{result.get('rtt', '')}</td>
            </tr>
"""
        
        html_content += """
        </tbody>
    </table>
</body>
</html>
"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _export_as_txt(self, filepath, results):
        """Export results as plain text"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("NetTools Scan Report\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Results: {len(results)}\n\n")
            f.write("-" * 80 + "\n")
            f.write(f"{'IP Address':<20} {'Hostname':<30} {'Status':<10} {'RTT':<10}\n")
            f.write("-" * 80 + "\n")
            for result in results:
                f.write(f"{result.get('ip', ''):<20} {result.get('hostname', '-'):<30} {result.get('status', ''):<10} {result.get('rtt', ''):<10}\n")
    
    def _export_as_xml(self, filepath, results):
        """Export results as XML"""
        xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml_content += '<scan_results>\n'
        xml_content += f'  <metadata>\n'
        xml_content += f'    <timestamp>{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</timestamp>\n'
        xml_content += f'    <total_results>{len(results)}</total_results>\n'
        xml_content += f'  </metadata>\n'
        xml_content += '  <hosts>\n'
        for result in results:
            xml_content += '    <host>\n'
            xml_content += f'      <ip>{result.get("ip", "")}</ip>\n'
            xml_content += f'      <hostname>{result.get("hostname", "")}</hostname>\n'
            xml_content += f'      <status>{result.get("status", "")}</status>\n'
            xml_content += f'      <response_time>{result.get("rtt", "")}</response_time>\n'
            xml_content += '    </host>\n'
        xml_content += '  </hosts>\n'
        xml_content += '</scan_results>\n'
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(xml_content)
            self.after(2000, lambda: self.status_label.configure(text="Ready."))
    
    def copy_textbox_to_clipboard(self, textbox):
        """Copy textbox content to clipboard"""
        text = textbox.get("1.0", 'end-1c')
        if text:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.status_label.configure(text="Copied to clipboard!")
            self.after(2000, lambda: self.status_label.configure(text="Ready."))
    
    def on_window_resize(self, event):
        """Handle window resize for auto-scaling"""
        # Only process resize events for the main window
        if event.widget != self:
            return
        
        # Calculate scale factor based on window size
        new_width = event.width
        new_height = event.height
        
        # Calculate scale (min 0.75, max 1.8)
        width_scale = new_width / self.base_width
        height_scale = new_height / self.base_height
        new_scale = min(width_scale, height_scale)
        new_scale = max(0.75, min(1.8, new_scale))
        
        # Only update if scale changed significantly (avoid too many updates)
        if abs(new_scale - self.current_scale) > 0.08:
            self.current_scale = new_scale
            self.update_fonts_and_sizes()
    
    def update_fonts_and_sizes(self):
        """Update fonts and widget sizes based on current scale"""
        try:
            # Calculate scaled font sizes
            scaled_base = max(9, int(self.base_font_size * self.current_scale))
            scaled_title = max(18, int(self.title_font_size * self.current_scale))
            scaled_subtitle = max(10, int(self.subtitle_font_size * self.current_scale))
            scaled_label = max(10, int(self.label_font_size * self.current_scale))
            
            # Update header fonts
            for widget in self.winfo_children():
                self.update_widget_fonts(widget, scaled_base, scaled_label)
                
        except Exception as e:
            pass  # Silently ignore scaling errors
    
    def update_widget_fonts(self, widget, base_size, label_size):
        """Recursively update fonts for all widgets"""
        try:
            if isinstance(widget, ctk.CTkLabel):
                current_font = widget.cget("font")
                if current_font:
                    if isinstance(current_font, ctk.CTkFont):
                        if current_font.cget("size") >= 14:
                            # Large labels (titles)
                            widget.configure(font=ctk.CTkFont(size=label_size + 2, weight=current_font.cget("weight")))
                        else:
                            # Normal labels
                            widget.configure(font=ctk.CTkFont(size=label_size, weight=current_font.cget("weight")))
            elif isinstance(widget, (ctk.CTkEntry, ctk.CTkTextbox)):
                # Scale entry and textbox fonts
                widget.configure(font=ctk.CTkFont(size=base_size))
            elif isinstance(widget, ctk.CTkButton):
                # Scale button fonts
                widget.configure(font=ctk.CTkFont(size=base_size))
            
            # Recursively update children
            for child in widget.winfo_children():
                self.update_widget_fonts(child, base_size, label_size)
        except:
            pass
    
    # Old tab-based methods removed - now using sidebar navigation
    
    def show_cidr_history(self):
        """Show CIDR history dropdown"""
        recent_cidrs = self.history.get_recent_cidrs()
        
        if not recent_cidrs:
            messagebox.showinfo("History", "No recent scans in history.")
            return
        
        # Create popup window
        history_window = ctk.CTkToplevel(self)
        history_window.title("Recent Scans")
        history_window.geometry("400x350")
        history_window.transient(self)
        history_window.grab_set()
        
        # Center window
        history_window.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - history_window.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - history_window.winfo_height()) // 2
        history_window.geometry(f"+{x}+{y}")
        
        # Ensure dialog is on top and focused
        history_window.lift()
        history_window.focus_force()
        
        # Title
        title_label = ctk.CTkLabel(
            history_window,
            text="Recent Scans (Click to use)",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title_label.pack(padx=20, pady=(20, 10))
        
        # Scrollable frame for history items
        scroll_frame = ctk.CTkScrollableFrame(history_window)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Add each history item as a button
        for cidr in recent_cidrs:
            btn = ctk.CTkButton(
                scroll_frame,
                text=cidr,
                command=lambda c=cidr: self.select_cidr_from_history(c, history_window),
                height=35,
                anchor="w"
            )
            btn.pack(fill="x", pady=3)
        
        # Bottom buttons
        button_frame = ctk.CTkFrame(history_window, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        clear_btn = ctk.CTkButton(
            button_frame,
            text="Clear History",
            command=lambda: self.clear_cidr_history(history_window),
            fg_color="#dc3545",
            hover_color="#c82333"
        )
        clear_btn.pack(side="left", padx=(0, 10))
        
        close_btn = ctk.CTkButton(
            button_frame,
            text="Close",
            command=history_window.destroy
        )
        close_btn.pack(side="left")
    
    def select_cidr_from_history(self, cidr, window):
        """Select CIDR from history"""
        self.cidr_entry.delete(0, 'end')
        self.cidr_entry.insert(0, cidr)
        self.update_host_count()
        window.destroy()
    
    def clear_cidr_history(self, window):
        """Clear CIDR history"""
        result = messagebox.askyesno(
            "Clear History",
            "Are you sure you want to clear all scan history?"
        )
        if result:
            self.history.clear_cidr_history()
            window.destroy()
            messagebox.showinfo("History", "Scan history cleared.")
    
    def show_mac_history(self):
        """Show MAC history dropdown"""
        recent_macs = self.history.get_recent_macs()
        
        if not recent_macs:
            messagebox.showinfo("History", "No recent MAC addresses in history.")
            return
        
        # Create popup window
        history_window = ctk.CTkToplevel(self)
        history_window.title("Recent MAC Addresses")
        history_window.geometry("450x350")
        history_window.transient(self)
        history_window.grab_set()
        
        # Center window
        history_window.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - history_window.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - history_window.winfo_height()) // 2
        history_window.geometry(f"+{x}+{y}")
        
        # Ensure dialog is on top and focused
        history_window.lift()
        history_window.focus_force()
        
        # Title
        title_label = ctk.CTkLabel(
            history_window,
            text="Recent MAC Addresses (Click to use)",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        title_label.pack(padx=20, pady=(20, 10))
        
        # Scrollable frame for history items
        scroll_frame = ctk.CTkScrollableFrame(history_window)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Add each history item as a button
        for mac in recent_macs:
            btn = ctk.CTkButton(
                scroll_frame,
                text=mac,
                command=lambda m=mac: self.select_mac_from_history(m, history_window),
                height=35,
                anchor="w"
            )
            btn.pack(fill="x", pady=3)
        
        # Bottom buttons
        button_frame = ctk.CTkFrame(history_window, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        clear_btn = ctk.CTkButton(
            button_frame,
            text="Clear History",
            command=lambda: self.clear_mac_history(history_window),
            fg_color="#dc3545",
            hover_color="#c82333"
        )
        clear_btn.pack(side="left", padx=(0, 10))
        
        close_btn = ctk.CTkButton(
            button_frame,
            text="Close",
            command=history_window.destroy
        )
        close_btn.pack(side="left")
    
    def select_mac_from_history(self, mac, window):
        """Select MAC from history"""
        self.mac_entry.delete(0, 'end')
        self.mac_entry.insert(0, mac)
        self.update_mac_formats()
        window.destroy()
    
    def clear_mac_history(self, window):
        """Clear MAC history"""
        result = messagebox.askyesno(
            "Clear History",
            "Are you sure you want to clear all MAC history?"
        )
        if result:
            self.history.clear_mac_history()
            window.destroy()
            messagebox.showinfo("History", "MAC history cleared.")
    
    def refresh_profiles(self):
        """Refresh the profiles list display"""
        # Clear existing
        for widget in self.profiles_frame.winfo_children():
            widget.destroy()
        
        profiles = self.profile_manager.profiles
        
        if not profiles:
            no_profiles_label = ctk.CTkLabel(
                self.profiles_frame,
                text="No saved profiles. Create one to get started!",
                font=ctk.CTkFont(size=12)
            )
            no_profiles_label.pack(pady=20)
            return
        
        # Display profiles in reverse order (newest first)
        for profile in reversed(profiles):
            self.create_profile_card(profile)
    
    def create_profile_card(self, profile):
        """Create a card for a saved profile"""
        card = ctk.CTkFrame(self.profiles_frame, corner_radius=8)
        card.pack(fill="x", pady=5)
        
        content_frame = ctk.CTkFrame(card, fg_color="transparent")
        content_frame.pack(fill="x", padx=15, pady=10)
        
        # Profile name
        name_label = ctk.CTkLabel(
            content_frame,
            text=profile["name"],
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="w"
        )
        name_label.pack(side="left", fill="x", expand=True)
        
        # Apply button
        apply_btn = ctk.CTkButton(
            content_frame,
            text="Apply Profile",
            command=lambda p=profile: self.apply_profile(p),
            width=120,
            height=36,
            fg_color=("#2196F3", "#1976D2")
        )
        apply_btn.pack(side="right", padx=(5, 0))
        
        # Delete button
        delete_btn = ctk.CTkButton(
            content_frame,
            text="Delete",
            command=lambda p=profile: self.delete_profile_confirm(p),
            width=80,
            height=36,
            fg_color=("#F44336", "#D32F2F")
        )
        delete_btn.pack(side="right", padx=(5, 0))
        
        # Show interface count
        interface_count = len(profile.get("interfaces", []))
        count_label = ctk.CTkLabel(
            card,
            text=f"{interface_count} interface(s) configured",
            font=ctk.CTkFont(size=11),
            text_color=("gray60", "gray40")
        )
        count_label.pack(padx=15, pady=(0, 10), anchor="w")
    
    def set_dhcp(self, interface_name):
        """Set interface to DHCP"""
        if not self.is_admin():
            # Try to elevate privileges via UAC
            result = messagebox.askyesno(
                "Administrator Required",
                "Administrator privileges are required to change network settings.\n\n"
                "Do you want to restart the application with administrator privileges?"
            )
            if result:
                self.restart_as_admin()
            return
        
        try:
            # Set DHCP for IP
            result = self.run_subprocess(
                ["netsh", "interface", "ipv4", "set", "address", interface_name, "dhcp"],
                capture_output=True,
                text=True,
                timeout=10
            )
            
            if result.returncode != 0:
                error_msg = result.stderr or result.stdout
                if "disconnected" in error_msg.lower():
                    messagebox.showwarning(
                        "Interface Disconnected",
                        f"Cannot configure '{interface_name}' because it is currently disconnected.\n\n"
                        "Please connect the network cable or enable Wi-Fi, then try again."
                    )
                else:
                    messagebox.showerror("Error", f"Failed to set DHCP:\n{error_msg}")
                return
            
            # Set DHCP for DNS
            self.run_subprocess(
                ["netsh", "interface", "ipv4", "set", "dnsservers", interface_name, "dhcp"],
                timeout=10
            )
            
            messagebox.showinfo("Success", f"Interface '{interface_name}' set to DHCP successfully!")
            self.refresh_interfaces()
            
        except subprocess.TimeoutExpired:
            messagebox.showerror("Timeout", "The command took too long to execute. Please try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
    
    def show_static_ip_dialog(self, interface_name):
        """Show dialog to configure static IP"""
        # Get current config
        current_config = self.get_interface_config(interface_name)
        
        # Create dialog window
        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Configure Static IP - {interface_name}")
        dialog.geometry("550x600")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center window
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - dialog.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Ensure dialog is on top and focused
        dialog.lift()
        dialog.focus_force()
        
        # Title section (fixed at top)
        title_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=(20, 0))
        
        title_label = ctk.CTkLabel(
            title_frame,
            text=f"Static IP Configuration",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(pady=(0, 5))
        
        subtitle_label = ctk.CTkLabel(
            title_frame,
            text=interface_name,
            font=ctk.CTkFont(size=12)
        )
        subtitle_label.pack(pady=(0, 15))
        
        # Scrollable form frame
        form_scroll = ctk.CTkScrollableFrame(dialog, height=380)
        form_scroll.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        # Inner form frame for consistent padding
        form_frame = ctk.CTkFrame(form_scroll, fg_color="transparent")
        form_frame.pack(fill="both", expand=True)
        
        # IP Address
        ip_label = ctk.CTkLabel(form_frame, text="IP Address:", font=ctk.CTkFont(size=12, weight="bold"))
        ip_label.pack(pady=(15, 5), anchor="w", padx=15)
        
        ip_entry = ctk.CTkEntry(form_frame, placeholder_text="e.g., 192.168.1.100", height=40)
        ip_entry.pack(fill="x", padx=15)
        if current_config and current_config["ip"]:
            ip_entry.insert(0, current_config["ip"])
        
        # Subnet Mask
        subnet_label = ctk.CTkLabel(form_frame, text="Subnet Mask:", font=ctk.CTkFont(size=12, weight="bold"))
        subnet_label.pack(pady=(15, 5), anchor="w", padx=15)
        
        subnet_entry = ctk.CTkEntry(form_frame, placeholder_text="e.g., 255.255.255.0", height=40)
        subnet_entry.pack(fill="x", padx=15)
        if current_config and current_config["subnet"]:
            subnet_entry.insert(0, current_config["subnet"])
        else:
            subnet_entry.insert(0, "255.255.255.0")
        
        # Default Gateway
        gateway_label = ctk.CTkLabel(form_frame, text="Default Gateway:", font=ctk.CTkFont(size=12, weight="bold"))
        gateway_label.pack(pady=(15, 5), anchor="w", padx=15)
        
        gateway_entry = ctk.CTkEntry(form_frame, placeholder_text="e.g., 192.168.1.1", height=40)
        gateway_entry.pack(fill="x", padx=15)
        if current_config and current_config["gateway"]:
            gateway_entry.insert(0, current_config["gateway"])
        
        # DNS Server (Primary)
        dns_label = ctk.CTkLabel(form_frame, text="Primary DNS:", font=ctk.CTkFont(size=12, weight="bold"))
        dns_label.pack(pady=(15, 5), anchor="w", padx=15)
        
        dns_entry = ctk.CTkEntry(form_frame, placeholder_text="e.g., 8.8.8.8 (optional)", height=40)
        dns_entry.pack(fill="x", padx=15, pady=(0, 15))
        if current_config and current_config["dns"] and len(current_config["dns"]) > 0:
            dns_entry.insert(0, current_config["dns"][0])
        
        # Button frame (fixed at bottom)
        button_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        button_frame.pack(fill="x", side="bottom", padx=20, pady=20)
        
        def apply_static():
            ip = ip_entry.get().strip()
            subnet = subnet_entry.get().strip()
            gateway = gateway_entry.get().strip()
            dns = dns_entry.get().strip()
            
            if not ip:
                messagebox.showwarning("Invalid Input", "IP Address is required")
                return
            
            if not subnet:
                messagebox.showwarning("Invalid Input", "Subnet Mask is required")
                return
            
            self.set_static_ip(interface_name, ip, subnet, gateway, dns)
            dialog.destroy()
        
        apply_btn = ctk.CTkButton(
            button_frame,
            text="Apply",
            command=apply_static,
            width=120,
            height=40,
            fg_color=("#4CAF50", "#388E3C")
        )
        apply_btn.pack(side="left", padx=(0, 10))
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            width=120,
            height=40
        )
        cancel_btn.pack(side="left")
    
    def set_static_ip(self, interface_name, ip, subnet, gateway, dns):
        """Set static IP configuration"""
        if not self.is_admin():
            result = messagebox.askyesno(
                "Administrator Required",
                "Administrator privileges are required to change network settings.\n\n"
                "Do you want to restart the application with administrator privileges?"
            )
            if result:
                self.restart_as_admin()
            return
        
        try:
            # Build command
            if gateway:
                cmd = ["netsh", "interface", "ipv4", "set", "address", interface_name, "static", ip, subnet, gateway]
            else:
                cmd = ["netsh", "interface", "ipv4", "set", "address", interface_name, "static", ip, subnet]
            
            result = self.run_subprocess(cmd, capture_output=True, text=True, timeout=10)
            
            if result.returncode != 0:
                error_msg = result.stderr or result.stdout
                if "disconnected" in error_msg.lower():
                    messagebox.showwarning(
                        "Interface Disconnected",
                        f"Cannot configure '{interface_name}' because it is currently disconnected.\n\n"
                        "Please connect the network cable or enable Wi-Fi, then try again."
                    )
                elif "not valid" in error_msg.lower() or "incorrect" in error_msg.lower():
                    messagebox.showerror(
                        "Invalid Configuration",
                        f"The IP configuration is invalid:\n\n{error_msg}\n\n"
                        "Please check:\n"
                        "• IP address format (e.g., 192.168.1.100)\n"
                        "• Subnet mask format (e.g., 255.255.255.0)\n"
                        "• Gateway is in the same subnet"
                    )
                else:
                    messagebox.showerror("Error", f"Failed to set static IP:\n{error_msg}")
                return
            
            # Set DNS if provided
            if dns:
                self.run_subprocess(
                    ["netsh", "interface", "ipv4", "set", "dnsservers", interface_name, "static", dns, "primary"],
                    timeout=10
                )
            
            messagebox.showinfo("Success", f"Static IP configured successfully!\n\nIP: {ip}\nSubnet: {subnet}")
            self.refresh_interfaces()
            
        except subprocess.TimeoutExpired:
            messagebox.showerror("Timeout", "The command took too long to execute. Please try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
    
    def create_new_profile(self):
        """Show dialog to create and configure a new profile"""
        # Get current interfaces
        interfaces = self.get_network_interfaces()
        
        if not interfaces:
            messagebox.showerror("Error", "No network interfaces found.")
            return
        
        # Create dialog window
        dialog = ctk.CTkToplevel(self)
        dialog.title("Create Network Profile")
        dialog.geometry("800x700")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center window
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - dialog.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Ensure dialog is on top and focused
        dialog.lift()
        dialog.focus_force()
        
        # Title
        title_label = ctk.CTkLabel(
            dialog,
            text="Create Network Profile",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(padx=20, pady=(20, 10))
        
        # Profile name
        name_frame = ctk.CTkFrame(dialog)
        name_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        name_label = ctk.CTkLabel(
            name_frame,
            text="Profile Name:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        name_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        name_entry = ctk.CTkEntry(
            name_frame,
            placeholder_text="e.g., Home Network, Office WiFi",
            height=40,
            font=ctk.CTkFont(size=13)
        )
        name_entry.pack(fill="x", padx=10, pady=(0, 10))
        
        # Instructions
        info_label = ctk.CTkLabel(
            dialog,
            text="Select interfaces to configure and set their network settings:",
            font=ctk.CTkFont(size=11),
            text_color=("gray60", "gray40")
        )
        info_label.pack(padx=20, pady=(0, 10))
        
        # Scrollable frame for interfaces
        scroll_frame = ctk.CTkScrollableFrame(dialog, height=400)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        # Store interface configurations
        interface_configs = {}
        
        # Display configurable interfaces
        for interface in interfaces:
            interface_frame = ctk.CTkFrame(scroll_frame, corner_radius=8)
            interface_frame.pack(fill="x", pady=5)
            
            # Checkbox to include this interface
            include_var = ctk.BooleanVar(value=True)
            include_check = ctk.CTkCheckBox(
                interface_frame,
                text=f"📶 {interface['name']}",
                variable=include_var,
                font=ctk.CTkFont(size=12, weight="bold")
            )
            include_check.pack(anchor="w", padx=10, pady=(10, 5))
            
            # Configuration section
            config_frame = ctk.CTkFrame(interface_frame, fg_color="transparent")
            config_frame.pack(fill="x", padx=20, pady=(0, 10))
            
            # DHCP or Static
            dhcp_var = ctk.StringVar(value="dhcp")
            
            # Function to toggle static fields (defined before widgets)
            def toggle_static_fields(*args):
                if dhcp_var.get() == "static":
                    ip_entry.configure(state="normal")
                    mask_entry.configure(state="normal")
                    gateway_entry.configure(state="normal")
                    dns_entry.configure(state="normal")
                else:
                    ip_entry.configure(state="disabled")
                    mask_entry.configure(state="disabled")
                    gateway_entry.configure(state="disabled")
                    dns_entry.configure(state="disabled")
            
            # Radio buttons at the top
            dhcp_radio = ctk.CTkRadioButton(
                config_frame,
                text="DHCP (Automatic)",
                variable=dhcp_var,
                value="dhcp",
                command=toggle_static_fields
            )
            dhcp_radio.pack(anchor="w", pady=2)
            
            static_radio = ctk.CTkRadioButton(
                config_frame,
                text="Static (Manual)",
                variable=dhcp_var,
                value="static",
                command=toggle_static_fields
            )
            static_radio.pack(anchor="w", pady=2)
            
            # Static IP fields frame
            static_frame = ctk.CTkFrame(config_frame, fg_color="transparent")
            static_frame.pack(fill="x", pady=(10, 0))
            
            # IP Address (Required)
            ip_label = ctk.CTkLabel(static_frame, text="IP Address: *", font=ctk.CTkFont(size=11))
            ip_label.grid(row=0, column=0, sticky="w", pady=2)
            ip_entry = ctk.CTkEntry(static_frame, placeholder_text="192.168.1.100", width=150, state="disabled")
            ip_entry.grid(row=0, column=1, padx=(10, 0), pady=2)
            
            # Subnet Mask (Required)
            mask_label = ctk.CTkLabel(static_frame, text="Subnet Mask: *", font=ctk.CTkFont(size=11))
            mask_label.grid(row=1, column=0, sticky="w", pady=2)
            mask_entry = ctk.CTkEntry(static_frame, placeholder_text="255.255.255.0", width=150, state="disabled")
            mask_entry.grid(row=1, column=1, padx=(10, 0), pady=2)
            
            # Gateway (Optional)
            gateway_label = ctk.CTkLabel(static_frame, text="Gateway (Optional):", font=ctk.CTkFont(size=11))
            gateway_label.grid(row=2, column=0, sticky="w", pady=2)
            gateway_entry = ctk.CTkEntry(static_frame, placeholder_text="192.168.1.1", width=150, state="disabled")
            gateway_entry.grid(row=2, column=1, padx=(10, 0), pady=2)
            
            # DNS (Optional)
            dns_label = ctk.CTkLabel(static_frame, text="DNS Server (Optional):", font=ctk.CTkFont(size=11))
            dns_label.grid(row=3, column=0, sticky="w", pady=2)
            dns_entry = ctk.CTkEntry(static_frame, placeholder_text="8.8.8.8, 8.8.4.4", width=150, state="disabled")
            dns_entry.grid(row=3, column=1, padx=(10, 0), pady=2)
            
            # Help text
            help_label = ctk.CTkLabel(
                static_frame,
                text="* Required fields. DNS supports multiple servers (comma-separated).",
                font=ctk.CTkFont(size=10),
                text_color=("gray60", "gray40")
            )
            help_label.grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))
            
            # Store references
            interface_configs[interface['name']] = {
                'include': include_var,
                'dhcp_var': dhcp_var,
                'ip_entry': ip_entry,
                'mask_entry': mask_entry,
                'gateway_entry': gateway_entry,
                'dns_entry': dns_entry
            }
        
        # Button frame
        button_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        def save_profile():
            """Save the configured profile"""
            name = name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Please enter a profile name.")
                return
            
            # Collect selected interface configurations
            saved_interfaces = []
            for interface_name, config_refs in interface_configs.items():
                if not config_refs['include'].get():
                    continue  # Skip unchecked interfaces
                
                # Build configuration
                if config_refs['dhcp_var'].get() == "dhcp":
                    config = {
                        "dhcp": True,
                        "ip": None,
                        "subnet": None,
                        "gateway": None,
                        "dns": []
                    }
                else:
                    ip = config_refs['ip_entry'].get().strip()
                    mask = config_refs['mask_entry'].get().strip()
                    gateway = config_refs['gateway_entry'].get().strip()
                    dns_input = config_refs['dns_entry'].get().strip()
                    
                    if not ip or not mask:
                        messagebox.showwarning("Warning", f"Static IP requires IP address and subnet mask for {interface_name}")
                        return
                    
                    # Parse DNS servers (support comma-separated list)
                    dns_servers = []
                    if dns_input:
                        dns_servers = [d.strip() for d in dns_input.split(',') if d.strip()]
                    
                    config = {
                        "dhcp": False,
                        "ip": ip,
                        "subnet": mask,
                        "subnet_mask": mask,
                        "gateway": gateway if gateway else None,
                        "dns": dns_servers
                    }
                
                saved_interfaces.append({
                    "name": interface_name,
                    "config": config
                })
            
            if not saved_interfaces:
                messagebox.showwarning("Warning", "Please select at least one interface to configure.")
                return
            
            # Save profile
            self.profile_manager.add_profile(name, saved_interfaces)
            self.refresh_profiles()
            
            dialog.destroy()
            messagebox.showinfo("Success", f"Profile '{name}' created with {len(saved_interfaces)} interface(s)!")
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="💾 Create Profile",
            command=save_profile,
            width=160,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=COLORS["success"],
            hover_color=COLORS["success_hover"]
        )
        save_btn.pack(side="right", padx=(10, 0))
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            width=120,
            height=40,
            font=ctk.CTkFont(size=14)
        )
        cancel_btn.pack(side="right")
    
    def apply_profile(self, profile):
        """Apply a saved profile"""
        if not self.is_admin():
            # Offer to restart as admin
            result = messagebox.askyesno(
                "Admin Required",
                "Administrator privileges are required to apply profiles.\n\nWould you like to restart the application as administrator?"
            )
            if result:
                self.restart_as_admin()
            return
        
        # Confirmation dialog
        result = messagebox.askyesno(
            "Apply Profile",
            f"Are you sure you want to apply the profile '{profile['name']}'?\n\nThis will change network settings for all configured interfaces."
        )
        
        if not result:
            return
        
        # Show progress message
        progress_label = ctk.CTkLabel(
            self.profiles_frame.master,
            text=f"Applying profile '{profile['name']}'...",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["primary"]
        )
        progress_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Apply profile in background thread
        def apply_in_background():
            try:
                success_count = 0
                error_count = 0
                errors = []
                
                for interface_data in profile["interfaces"]:
                    interface_name = interface_data["name"]
                    config = interface_data["config"]
                    
                    try:
                        # Check if interface still exists
                        current_interfaces = self.get_network_interfaces()
                        if not any(i["name"] == interface_name for i in current_interfaces):
                            errors.append(f"{interface_name}: Interface not found")
                            error_count += 1
                            continue
                        
                        # Apply configuration
                        if config["dhcp"]:
                            # Set to DHCP
                            result = self.run_subprocess(
                                ["netsh", "interface", "ipv4", "set", "address", 
                                 interface_name, "dhcp"],
                                capture_output=True,
                                text=True,
                                timeout=10
                            )
                            
                            if result.returncode == 0:
                                # Set DNS to DHCP
                                self.run_subprocess(
                                    ["netsh", "interface", "ipv4", "set", "dnsservers",
                                     interface_name, "dhcp"],
                                    capture_output=True,
                                    text=True,
                                    timeout=10
                                )
                                success_count += 1
                            else:
                                errors.append(f"{interface_name}: Failed to set DHCP")
                                error_count += 1
                        else:
                            # Set static IP
                            if config["ip"] and config["subnet"]:
                                # Determine subnet mask from prefix or use default
                                subnet_mask = config.get("subnet_mask", config.get("subnet", "255.255.255.0"))
                                gateway = config.get("gateway", "none")
                                
                                # Build netsh command for static IP
                                cmd = [
                                    "netsh", "interface", "ipv4", "set", "address",
                                    "name=" + interface_name,
                                    "source=static",
                                    "address=" + config["ip"],
                                    "mask=" + subnet_mask,
                                    "gateway=" + gateway
                                ]
                                
                                result = self.run_subprocess(
                                    cmd,
                                    capture_output=True,
                                    text=True,
                                    encoding='cp850',
                                    timeout=15
                                )
                                
                                if result.returncode == 0:
                                    # Set DNS if configured
                                    if config.get("dns") and len(config["dns"]) > 0:
                                        for i, dns in enumerate(config["dns"]):
                                            if dns.strip():  # Only add non-empty DNS
                                                if i == 0:
                                                    # Primary DNS
                                                    dns_cmd = [
                                                        "netsh", "interface", "ipv4", "set", "dns",
                                                        "name=" + interface_name,
                                                        "source=static",
                                                        "address=" + dns,
                                                        "register=primary"
                                                    ]
                                                else:
                                                    # Additional DNS servers
                                                    dns_cmd = [
                                                        "netsh", "interface", "ipv4", "add", "dns",
                                                        "name=" + interface_name,
                                                        "address=" + dns,
                                                        "index=" + str(i + 1)
                                                    ]
                                                self.run_subprocess(dns_cmd, capture_output=True, timeout=10)
                                    
                                    success_count += 1
                                else:
                                    error_msg = result.stderr if result.stderr else "Unknown error"
                                    errors.append(f"{interface_name}: {error_msg[:100]}")
                                    error_count += 1
                            else:
                                errors.append(f"{interface_name}: Missing IP or subnet mask")
                                error_count += 1
                    
                    except Exception as e:
                        errors.append(f"{interface_name}: {str(e)}")
                        error_count += 1
                
                # Hide progress and show results
                self.after(0, lambda: progress_label.destroy())
                self.after(0, self.refresh_interfaces)
                
                if error_count == 0:
                    self.after(0, lambda: messagebox.showinfo(
                        "Success",
                        f"Profile '{profile['name']}' applied successfully!\n\n{success_count} interface(s) configured."
                    ))
                else:
                    error_msg = f"Profile partially applied.\n\nSuccessful: {success_count}\nFailed: {error_count}"
                    if errors:
                        error_msg += f"\n\nErrors:\n" + "\n".join(errors[:5])
                    self.after(0, lambda: messagebox.showwarning(
                        "Partial Success",
                        error_msg
                    ))
            
            except Exception as e:
                self.after(0, lambda: progress_label.destroy())
                self.after(0, lambda: messagebox.showerror(
                    "Error",
                    f"Error applying profile: {str(e)}"
                ))
        
        # Run in thread
        thread = threading.Thread(target=apply_in_background, daemon=True)
        thread.start()
    
    def delete_profile_confirm(self, profile):
        """Confirm and delete a profile"""
        result = messagebox.askyesno(
            "Delete Profile",
            f"Are you sure you want to delete the profile '{profile['name']}'?\n\nThis cannot be undone."
        )
        
        if result:
            self.profile_manager.delete_profile(profile["id"])
            self.refresh_profiles()
            messagebox.showinfo("Success", f"Profile '{profile['name']}' deleted successfully.")
    
    def show_scan_comparison(self):
        """Show scan comparison window"""
        scans = self.scan_manager.get_scans()
        
        if len(scans) < 2:
            messagebox.showinfo("Scan Comparison", "You need at least 2 saved scans to compare. Run more scans first.")
            return
        
        # Create comparison window
        comp_window = ctk.CTkToplevel(self)
        comp_window.title("Scan Comparison")
        comp_window.geometry("900x700")
        comp_window.transient(self)
        comp_window.grab_set()
        
        # Center window
        comp_window.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - comp_window.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - comp_window.winfo_height()) // 2
        comp_window.geometry(f"+{x}+{y}")
        
        # Ensure dialog is on top and focused
        comp_window.lift()
        comp_window.focus_force()
        
        # Title
        title_label = ctk.CTkLabel(
            comp_window,
            text="Compare Network Scans",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(padx=20, pady=(20, 10))
        
        # Selection frame
        select_frame = ctk.CTkFrame(comp_window)
        select_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        # Prepare scan options
        scan_options = [f"{s['id']} - {s['cidr']} ({s['summary']['online']}/{s['summary']['total']} online)" for s in scans]
        
        # Scan 1 selection
        scan1_label = ctk.CTkLabel(select_frame, text="Scan 1:", font=ctk.CTkFont(size=12, weight="bold"))
        scan1_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        scan1_var = ctk.StringVar(value=scan_options[0])
        scan1_menu = ctk.CTkOptionMenu(select_frame, variable=scan1_var, values=scan_options, width=400)
        scan1_menu.grid(row=0, column=1, padx=10, pady=10)
        
        # Scan 2 selection
        scan2_label = ctk.CTkLabel(select_frame, text="Scan 2:", font=ctk.CTkFont(size=12, weight="bold"))
        scan2_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        
        scan2_var = ctk.StringVar(value=scan_options[-1] if len(scan_options) > 1 else scan_options[0])
        scan2_menu = ctk.CTkOptionMenu(select_frame, variable=scan2_var, values=scan_options, width=400)
        scan2_menu.grid(row=1, column=1, padx=10, pady=10)
        
        # Results area
        results_frame = ctk.CTkFrame(comp_window)
        results_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        results_scroll = ctk.CTkScrollableFrame(results_frame)
        results_scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        results_label = ctk.CTkLabel(
            results_scroll,
            text="Select two scans and click 'Compare' to see differences",
            font=ctk.CTkFont(size=12)
        )
        results_label.pack(pady=50)
        
        def do_comparison():
            """Perform the comparison"""
            # Get selected scan IDs
            scan1_id = scan1_var.get().split(" - ")[0]
            scan2_id = scan2_var.get().split(" - ")[0]
            
            if scan1_id == scan2_id:
                messagebox.showwarning("Same Scan", "Please select two different scans to compare.")
                return
            
            # Perform comparison
            comparison_result = self.scan_manager.compare_scans(scan1_id, scan2_id)
            
            if not comparison_result:
                messagebox.showerror("Error", "Could not load scan data.")
                return
            
            # Clear results
            for widget in results_scroll.winfo_children():
                widget.destroy()
            
            # Summary
            summary_frame = ctk.CTkFrame(results_scroll)
            summary_frame.pack(fill="x", padx=5, pady=10)
            
            summary_title = ctk.CTkLabel(
                summary_frame,
                text="Summary",
                font=ctk.CTkFont(size=14, weight="bold")
            )
            summary_title.pack(pady=(10, 5))
            
            summary = comparison_result["summary"]
            summary_text = f"✅ Unchanged: {summary['unchanged']}  |  "
            summary_text += f"🆕 New: {summary['new']}  |  "
            summary_text += f"❌ Missing: {summary['missing']}  |  "
            summary_text += f"🔄 Changed: {summary['changed']}"
            
            summary_label = ctk.CTkLabel(
                summary_frame,
                text=summary_text,
                font=ctk.CTkFont(size=12)
            )
            summary_label.pack(pady=(0, 10))
            
            # Details header
            details_header = ctk.CTkFrame(results_scroll)
            details_header.pack(fill="x", padx=5, pady=5)
            
            headers = [
                ("Change", 100),
                ("IP Address", 150),
                ("Scan 1 Status", 150),
                ("Scan 2 Status", 150),
                ("Scan 1 RTT", 120),
                ("Scan 2 RTT", 120)
            ]
            
            for text, width in headers:
                label = ctk.CTkLabel(
                    details_header,
                    text=text,
                    font=ctk.CTkFont(size=11, weight="bold"),
                    width=width
                )
                label.pack(side="left", padx=5)
            
            # Details rows
            for item in comparison_result["comparison"]:
                # Skip unchanged if there are too many
                if item["change"] == "unchanged" and summary["unchanged"] > 50:
                    # Only show first 10 unchanged
                    continue
                
                row_frame = ctk.CTkFrame(results_scroll)
                row_frame.pack(fill="x", padx=5, pady=2)
                
                # Change indicator
                change_icons = {
                    "unchanged": "✅",
                    "new": "🆕",
                    "missing": "❌",
                    "changed": "🔄"
                }
                change_colors = {
                    "unchanged": "#4CAF50",
                    "new": "#2196F3",
                    "missing": "#dc3545",
                    "changed": "#FF9800"
                }
                
                change_label = ctk.CTkLabel(
                    row_frame,
                    text=change_icons.get(item["change"], "?"),
                    width=100,
                    text_color=change_colors.get(item["change"], "#FFFFFF")
                )
                change_label.pack(side="left", padx=5)
                
                # IP
                ip_label = ctk.CTkLabel(row_frame, text=item["ip"], width=150)
                ip_label.pack(side="left", padx=5)
                
                # Scan 1 Status
                s1_label = ctk.CTkLabel(row_frame, text=item["scan1_status"], width=150)
                s1_label.pack(side="left", padx=5)
                
                # Scan 2 Status
                s2_label = ctk.CTkLabel(row_frame, text=item["scan2_status"], width=150)
                s2_label.pack(side="left", padx=5)
                
                # Scan 1 RTT
                s1_rtt = item["scan1_rtt"] if item["scan1_rtt"] else "-"
                s1_rtt_label = ctk.CTkLabel(row_frame, text=s1_rtt, width=120)
                s1_rtt_label.pack(side="left", padx=5)
                
                # Scan 2 RTT
                s2_rtt = item["scan2_rtt"] if item["scan2_rtt"] else "-"
                s2_rtt_label = ctk.CTkLabel(row_frame, text=s2_rtt, width=120)
                s2_rtt_label.pack(side="left", padx=5)
        
        def export_comparison():
            """Export comparison to file"""
            scan1_id = scan1_var.get().split(" - ")[0]
            scan2_id = scan2_var.get().split(" - ")[0]
            
            if scan1_id == scan2_id:
                messagebox.showwarning("Same Scan", "Please select two different scans to compare.")
                return
            
            comparison_result = self.scan_manager.compare_scans(scan1_id, scan2_id)
            if not comparison_result:
                messagebox.showerror("Error", "Could not load scan data.")
                return
            
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=f"comparison_{scan1_id}_vs_{scan2_id}.csv"
            )
            
            if filename:
                try:
                    with open(filename, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(['Change', 'IP Address', 'Scan 1 Status', 'Scan 2 Status', 'Scan 1 RTT', 'Scan 2 RTT'])
                        
                        for item in comparison_result["comparison"]:
                            writer.writerow([
                                item["change"],
                                item["ip"],
                                item["scan1_status"],
                                item["scan2_status"],
                                item["scan1_rtt"] or "-",
                                item["scan2_rtt"] or "-"
                            ])
                    
                    messagebox.showinfo("Success", f"Comparison exported to:\n{filename}")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not export comparison:\n{str(e)}")
        
        # Action buttons
        button_frame = ctk.CTkFrame(comp_window)
        button_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        compare_btn = ctk.CTkButton(
            button_frame,
            text="Compare",
            command=do_comparison,
            width=150
        )
        compare_btn.pack(side="left", padx=(0, 10))
        
        export_btn = ctk.CTkButton(
            button_frame,
            text="Export Comparison",
            command=export_comparison,
            width=150
        )
        export_btn.pack(side="left", padx=(0, 10))
        
        close_btn = ctk.CTkButton(
            button_frame,
            text="Close",
            command=comp_window.destroy,
            width=100
        )
        close_btn.pack(side="right")
    
    def create_phpipam_content(self, parent):
        """Create phpIPAM integration page content"""
        if not PHPIPAM_AVAILABLE:
            # Show error if modules not available
            error_frame = ctk.CTkFrame(parent)
            error_frame.pack(fill="both", expand=True, padx=20, pady=20)
            
            error_label = ctk.CTkLabel(
                error_frame,
                text="⚠️ phpIPAM Integration Unavailable",
                font=ctk.CTkFont(size=24, weight="bold")
            )
            error_label.pack(pady=(50, 10))
            
            msg_label = ctk.CTkLabel(
                error_frame,
                text="Required modules are missing. Please install:\n\npip install cryptography requests",
                font=ctk.CTkFont(size=14)
            )
            msg_label.pack(pady=20)
            return
        
        # Initialize phpIPAM config
        self.phpipam_config = PHPIPAMConfig()
        self.phpipam_client = None
        
        # Scrollable content area
        scrollable = ctk.CTkScrollableFrame(parent)
        scrollable.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = ctk.CTkLabel(
            scrollable,
            text="phpIPAM Integration",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=(0, 5))
        
        subtitle_label = SubTitle(
            scrollable,
            text="Manage IP addresses with phpIPAM API"
        )
        subtitle_label.pack(pady=(0, SPACING['lg']))
        
        # Status Section with styled card
        status_frame = StyledCard(scrollable)
        status_frame.pack(fill="x", pady=(0, SPACING['lg']))
        
        status_title = ctk.CTkLabel(
            status_frame,
            text="Connection Status",
            font=ctk.CTkFont(size=FONTS['subheading'], weight="bold")
        )
        status_title.pack(pady=(SPACING['lg'], SPACING['xs']), padx=SPACING['lg'], anchor="w")
        
        self.phpipam_status_label = ctk.CTkLabel(
            status_frame,
            text="⚪ Not configured" if not self.phpipam_config.is_enabled() else "🟢 Enabled",
            font=ctk.CTkFont(size=FONTS['body'])
        )
        self.phpipam_status_label.pack(pady=(0, SPACING['lg']), padx=SPACING['lg'], anchor="w")
        
        # Action buttons
        button_frame = ctk.CTkFrame(scrollable, fg_color="transparent")
        button_frame.pack(fill="x", pady=(0, SPACING['lg']))
        
        settings_btn = StyledButton(
            button_frame,
            text="⚙️ Settings",
            command=self.show_phpipam_settings,
            size="medium",
            variant="neutral"
        )
        settings_btn.pack(side="left", padx=(0, SPACING['md']))
        
        test_btn = StyledButton(
            button_frame,
            text="🔌 Test Connection",
            command=self.test_phpipam_connection,
            size="large",
            variant="primary"
        )
        test_btn.pack(side="left", padx=(0, SPACING['md']))
        
        auth_btn = StyledButton(
            button_frame,
            text="🔑 Authenticate",
            command=self.authenticate_phpipam,
            size="medium",
            variant="success"
        )
        auth_btn.pack(side="left")
        
        # Operations Section with styled card
        ops_frame = StyledCard(scrollable)
        ops_frame.pack(fill="x", pady=(0, SPACING['lg']))
        
        ops_title = SectionTitle(
            ops_frame,
            text="Operations"
        )
        ops_title.pack(pady=(SPACING['lg'], SPACING['md']), padx=SPACING['lg'], anchor="w")
        
        # IP Search
        search_frame = ctk.CTkFrame(ops_frame, fg_color="transparent")
        search_frame.pack(fill="x", padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        search_label = ctk.CTkLabel(
            search_frame,
            text="Search IP Address:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        search_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        search_entry_frame = ctk.CTkFrame(search_frame, fg_color="transparent")
        search_entry_frame.pack(fill="x")
        
        self.phpipam_search_entry = StyledEntry(
            search_entry_frame,
            placeholder_text="e.g., 192.168.1.10"
        )
        self.phpipam_search_entry.pack(side="left", fill="x", expand=True, padx=(0, SPACING['md']))
        
        search_btn = StyledButton(
            search_entry_frame,
            text="🔍 Search",
            command=self.search_phpipam_ip,
            size="medium",
            variant="primary"
        )
        search_btn.pack(side="left")
        
        # View Subnets
        subnet_btn_frame = ctk.CTkFrame(ops_frame, fg_color="transparent")
        subnet_btn_frame.pack(fill="x", padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        view_subnets_btn = StyledButton(
            subnet_btn_frame,
            text="📋 View All Subnets",
            command=self.view_phpipam_subnets,
            size="large",
            variant="primary"
        )
        view_subnets_btn.pack(side="left")
        
        # Results Section
        results_header_frame = ctk.CTkFrame(scrollable, fg_color="transparent")
        results_header_frame.pack(fill="x", pady=(SPACING['md'], SPACING['md']))
        
        results_title = SectionTitle(
            results_header_frame,
            text="Results"
        )
        results_title.pack(side="left", anchor="w")
        
        # Filter box for results
        filter_frame = ctk.CTkFrame(scrollable, corner_radius=6)
        filter_frame.pack(fill="x", pady=(0, 10))
        
        filter_label = ctk.CTkLabel(
            filter_frame,
            text="Filter results:",
            font=ctk.CTkFont(size=11)
        )
        filter_label.pack(side="left", padx=(10, 5))
        
        self.phpipam_filter_entry = ctk.CTkEntry(
            filter_frame,
            placeholder_text="Type to filter displayed results...",
            height=32,
            width=300
        )
        self.phpipam_filter_entry.pack(side="left", padx=5)
        self.phpipam_filter_entry.bind('<KeyRelease>', self._filter_phpipam_results)
        
        clear_filter_btn = ctk.CTkButton(
            filter_frame,
            text="✕",
            command=lambda: (self.phpipam_filter_entry.delete(0, 'end'), self._filter_phpipam_results()),
            width=30,
            height=32,
            fg_color=COLORS["neutral"],
            hover_color=COLORS["neutral_hover"]
        )
        clear_filter_btn.pack(side="left", padx=(0, 10))
        
        self.phpipam_results_frame = ctk.CTkFrame(scrollable, corner_radius=8)
        self.phpipam_results_frame.pack(fill="both", expand=True)
        
        # Initial message
        no_results_label = ctk.CTkLabel(
            self.phpipam_results_frame,
            text="No results yet. Configure settings and perform an operation.",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_secondary"]
        )
        no_results_label.pack(pady=50)
    
    def show_phpipam_settings(self):
        """Show phpIPAM settings dialog"""
        if not PHPIPAM_AVAILABLE:
            messagebox.showerror("Error", "phpIPAM modules not available")
            return
        
        # Create settings dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("phpIPAM Settings")
        dialog.geometry("600x700")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center window
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - dialog.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Ensure dialog is on top and focused
        dialog.lift()
        dialog.focus_force()
        
        # Content
        content = ctk.CTkScrollableFrame(dialog)
        content.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = ctk.CTkLabel(
            content,
            text="phpIPAM Configuration",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(pady=(0, 20))
        
        # Enable/Disable
        enabled_var = ctk.BooleanVar(value=self.phpipam_config.is_enabled())
        enabled_check = ctk.CTkCheckBox(
            content,
            text="Enable phpIPAM Integration",
            variable=enabled_var,
            font=ctk.CTkFont(size=13, weight="bold")
        )
        enabled_check.pack(anchor="w", pady=(0, 20))
        
        # phpIPAM URL
        url_label = ctk.CTkLabel(
            content,
            text="phpIPAM URL:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        url_label.pack(anchor="w", pady=(0, 5))
        
        url_entry = ctk.CTkEntry(
            content,
            placeholder_text="https://ipam.example.com",
            height=38
        )
        url_entry.insert(0, self.phpipam_config.get_phpipam_url())
        url_entry.pack(fill="x", pady=(0, 15))
        
        # App ID
        appid_label = ctk.CTkLabel(
            content,
            text="App ID:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        appid_label.pack(anchor="w", pady=(0, 5))
        
        appid_info = ctk.CTkLabel(
            content,
            text="Must match an API application created in phpIPAM (Administration → API)",
            font=ctk.CTkFont(size=10),
            text_color=COLORS["text_secondary"]
        )
        appid_info.pack(anchor="w", pady=(0, 5))
        
        appid_entry = ctk.CTkEntry(
            content,
            placeholder_text="MyApplication",
            height=38
        )
        appid_entry.insert(0, self.phpipam_config.get_app_id())
        appid_entry.pack(fill="x", pady=(0, 15))
        
        # Authentication Method
        auth_label = ctk.CTkLabel(
            content,
            text="Authentication Method:",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        auth_label.pack(anchor="w", pady=(0, 5))
        
        auth_var = ctk.StringVar(value=self.phpipam_config.get_auth_method())
        
        dynamic_radio = ctk.CTkRadioButton(
            content,
            text="Dynamic (Username + Password)",
            variable=auth_var,
            value="dynamic"
        )
        dynamic_radio.pack(anchor="w", pady=2)
        
        static_radio = ctk.CTkRadioButton(
            content,
            text="Static Token",
            variable=auth_var,
            value="static"
        )
        static_radio.pack(anchor="w", pady=(2, 15))
        
        # Username (for dynamic)
        username_label = ctk.CTkLabel(
            content,
            text="Username (Dynamic Auth):",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        username_label.pack(anchor="w", pady=(0, 5))
        
        username_entry = ctk.CTkEntry(
            content,
            placeholder_text="admin",
            height=38
        )
        username_entry.insert(0, self.phpipam_config.get_username())
        username_entry.pack(fill="x", pady=(0, 15))
        
        # Password (for dynamic)
        password_label = ctk.CTkLabel(
            content,
            text="Password (Dynamic Auth):",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        password_label.pack(anchor="w", pady=(0, 5))
        
        password_entry = ctk.CTkEntry(
            content,
            placeholder_text="Enter password",
            height=38,
            show="*"
        )
        password_entry.pack(fill="x", pady=(0, 15))
        
        # Static Token
        token_label = ctk.CTkLabel(
            content,
            text="Static Token (Static Auth):",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        token_label.pack(anchor="w", pady=(0, 5))
        
        token_entry = ctk.CTkEntry(
            content,
            placeholder_text="Enter static API token",
            height=38,
            show="*"
        )
        token_entry.pack(fill="x", pady=(0, 15))
        
        # SSL Verify
        ssl_var = ctk.BooleanVar(value=self.phpipam_config.get_ssl_verify())
        ssl_check = ctk.CTkCheckBox(
            content,
            text="Verify SSL Certificates",
            variable=ssl_var
        )
        ssl_check.pack(anchor="w", pady=(0, 20))
        
        # Save function
        def save_settings():
            password = password_entry.get()
            token = token_entry.get()
            
            # Don't overwrite if fields are empty (means user didn't change them)
            if not password:
                password = self.phpipam_config.get_password()
            if not token:
                token = self.phpipam_config.get_static_token()
            
            success = self.phpipam_config.update_config(
                enabled=enabled_var.get(),
                phpipam_url=url_entry.get(),
                app_id=appid_entry.get(),
                auth_method=auth_var.get(),
                username=username_entry.get(),
                password=password,
                static_token=token,
                ssl_verify=ssl_var.get()
            )
            
            if success:
                # Update status label
                if self.phpipam_config.is_enabled():
                    self.phpipam_status_label.configure(text="🟢 Enabled")
                else:
                    self.phpipam_status_label.configure(text="⚪ Disabled")
                
                # Reset client to force re-auth
                self.phpipam_client = None
                
                dialog.destroy()
                messagebox.showinfo("Success", "Settings saved successfully!")
            else:
                messagebox.showerror("Error", "Failed to save settings")
        
        # Buttons
        button_frame = ctk.CTkFrame(content, fg_color="transparent")
        button_frame.pack(fill="x", pady=(20, 0))
        
        save_btn = ctk.CTkButton(
            button_frame,
            text="💾 Save",
            command=save_settings,
            width=120,
            height=40,
            fg_color=COLORS["success"],
            hover_color=COLORS["success_hover"]
        )
        save_btn.pack(side="right", padx=(10, 0))
        
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            width=100,
            height=40
        )
        cancel_btn.pack(side="right")
    
    def test_phpipam_connection(self):
        """Test connection to phpIPAM"""
        if not PHPIPAM_AVAILABLE:
            messagebox.showerror("Error", "phpIPAM modules not available")
            return
        
        if not self.phpipam_config.is_enabled():
            messagebox.showwarning("Not Enabled", "phpIPAM integration is disabled. Enable it in Settings.")
            return
        
        # Reload config to get latest SSL settings
        self.phpipam_config.config = self.phpipam_config.load_config()
        
        # Create fresh client with updated config and test
        client = PHPIPAMClient(self.phpipam_config)
        success, message = client.test_connection()
        
        if success:
            messagebox.showinfo("Connection Test - Success", message)
        else:
            messagebox.showerror("Connection Test - Failed", message)
    
    def authenticate_phpipam(self):
        """Authenticate with phpIPAM"""
        if not PHPIPAM_AVAILABLE:
            messagebox.showerror("Error", "phpIPAM modules not available")
            return
        
        if not self.phpipam_config.is_enabled():
            messagebox.showwarning("Not Enabled", "phpIPAM integration is disabled. Enable it in Settings.")
            return
        
        # Create/get client
        if not self.phpipam_client:
            self.phpipam_client = PHPIPAMClient(self.phpipam_config)
        
        success, message = self.phpipam_client.authenticate()
        
        if success:
            messagebox.showinfo("Success", f"✅ {message}")
        else:
            messagebox.showerror("Authentication Failed", f"❌ {message}")
    
    def search_phpipam_ip(self):
        """Search for IP address in phpIPAM"""
        if not PHPIPAM_AVAILABLE:
            messagebox.showerror("Error", "phpIPAM modules not available")
            return
        
        if not self.phpipam_config.is_enabled():
            messagebox.showwarning("Not Enabled", "phpIPAM integration is disabled. Enable it in Settings.")
            return
        
        ip_address = self.phpipam_search_entry.get().strip()
        if not ip_address:
            messagebox.showwarning("Input Required", "Please enter an IP address to search")
            return
        
        # Show loading message
        self.display_phpipam_loading("Searching for IP address...")
        
        # Create/get client
        if not self.phpipam_client:
            self.phpipam_client = PHPIPAMClient(self.phpipam_config)
        
        # Run search in background thread
        def search_thread():
            success, results = self.phpipam_client.search_ip(ip_address)
            # Update UI in main thread
            self.after(0, self.display_phpipam_results, "IP Search Results", results, success)
        
        thread = threading.Thread(target=search_thread, daemon=True)
        thread.start()
    
    def view_phpipam_subnets(self):
        """View all subnets from phpIPAM"""
        if not PHPIPAM_AVAILABLE:
            messagebox.showerror("Error", "phpIPAM modules not available")
            return
        
        if not self.phpipam_config.is_enabled():
            messagebox.showwarning("Not Enabled", "phpIPAM integration is disabled. Enable it in Settings.")
            return
        
        # Show loading message
        self.display_phpipam_loading("Loading subnets from phpIPAM...")
        
        # Create/get client
        if not self.phpipam_client:
            self.phpipam_client = PHPIPAMClient(self.phpipam_config)
        
        # Get subnets in background thread
        def subnets_thread():
            success, results = self.phpipam_client.get_all_subnets()
            # Update UI in main thread
            self.after(0, self.display_phpipam_results, "All Subnets", results, success)
        
        thread = threading.Thread(target=subnets_thread, daemon=True)
        thread.start()
    
    def display_phpipam_loading(self, message):
        """Display loading message"""
        # Clear existing results
        for widget in self.phpipam_results_frame.winfo_children():
            widget.destroy()
        
        loading_label = ctk.CTkLabel(
            self.phpipam_results_frame,
            text=f"⏳ {message}",
            font=ctk.CTkFont(size=14),
            text_color=COLORS["primary"]
        )
        loading_label.pack(pady=50)
    
    def display_phpipam_results(self, title, data, success):
        """Display phpIPAM operation results with pagination for large datasets"""
        # Clear existing results
        for widget in self.phpipam_results_frame.winfo_children():
            widget.destroy()
        
        # Title
        result_title = ctk.CTkLabel(
            self.phpipam_results_frame,
            text=title,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        result_title.pack(pady=(15, 10), padx=15, anchor="w")
        
        if not success:
            error_label = ctk.CTkLabel(
                self.phpipam_results_frame,
                text=f"❌ Error: {data}",
                font=ctk.CTkFont(size=12),
                text_color=COLORS["danger"]
            )
            error_label.pack(pady=20, padx=15)
            return
        
        if not data:
            no_data_label = ctk.CTkLabel(
                self.phpipam_results_frame,
                text="No results found",
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_secondary"]
            )
            no_data_label.pack(pady=20, padx=15)
            return
        
        # Store data for pagination and filtering
        self.phpipam_all_results = data  # Keep original for filtering
        self.phpipam_current_results = data
        self.phpipam_current_page = 0
        self.phpipam_items_per_page = 50  # Show 50 items at a time
        
        # Info bar with count and warning for large datasets
        info_frame = ctk.CTkFrame(self.phpipam_results_frame, fg_color="transparent")
        info_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        count_text = f"Found {len(data)} result(s)"
        if len(data) > 100:
            count_text += f" • Showing {self.phpipam_items_per_page} per page for performance"
        
        count_label = ctk.CTkLabel(
            info_frame,
            text=count_text,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLORS["text_secondary"]
        )
        count_label.pack(side="left")
        
        # Scrollable results container
        self.phpipam_results_scroll = ctk.CTkScrollableFrame(
            self.phpipam_results_frame, 
            height=300
        )
        self.phpipam_results_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 10))
        
        # Display first page
        self._display_phpipam_page()
        
        # Pagination controls if needed
        if len(data) > self.phpipam_items_per_page:
            self._create_pagination_controls()
    
    def _create_subnet_card(self, parent, subnet_data):
        """Create formatted card for subnet display"""
        # Main subnet info
        subnet_str = subnet_data.get("subnet", "N/A")
        mask = subnet_data.get("mask", "")
        cidr = f"{subnet_str}/{mask}" if mask else subnet_str
        
        # Header with CIDR
        header_frame = ctk.CTkFrame(parent, fg_color="transparent")
        header_frame.pack(fill="x", padx=12, pady=(12, 5))
        
        cidr_label = ctk.CTkLabel(
            header_frame,
            text=f"🌐 {cidr}",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        cidr_label.pack(side="left")
        
        # ID badge
        subnet_id = subnet_data.get("id", "")
        if subnet_id:
            id_label = ctk.CTkLabel(
                header_frame,
                text=f"ID: {subnet_id}",
                font=ctk.CTkFont(size=10),
                text_color=COLORS["text_secondary"]
            )
            id_label.pack(side="right")
        
        # Description
        description = subnet_data.get("description", "")
        if description:
            desc_label = ctk.CTkLabel(
                parent,
                text=description,
                font=ctk.CTkFont(size=11),
                anchor="w",
                text_color=COLORS["text_secondary"]
            )
            desc_label.pack(fill="x", padx=12, pady=(0, 5))
        
        # Additional info in compact format
        info_items = []
        
        if "vlanId" in subnet_data and subnet_data["vlanId"]:
            info_items.append(f"VLAN: {subnet_data['vlanId']}")
        
        if "location" in subnet_data and subnet_data["location"]:
            info_items.append(f"Location: {subnet_data['location']}")
        
        if "isFolder" in subnet_data and subnet_data["isFolder"] == "1":
            info_items.append("📁 Folder")
        
        if info_items:
            info_text = " • ".join(info_items)
            info_label = ctk.CTkLabel(
                parent,
                text=info_text,
                font=ctk.CTkFont(size=10),
                anchor="w",
                text_color=COLORS["text_secondary"]
            )
            info_label.pack(fill="x", padx=12, pady=(0, 12))
        else:
            # Add bottom padding
            ctk.CTkFrame(parent, height=5, fg_color="transparent").pack()
    
    def _display_phpipam_page(self):
        """Display current page of phpIPAM results"""
        # Clear scroll frame
        for widget in self.phpipam_results_scroll.winfo_children():
            widget.destroy()
        
        # Calculate page boundaries
        start_idx = self.phpipam_current_page * self.phpipam_items_per_page
        end_idx = min(start_idx + self.phpipam_items_per_page, len(self.phpipam_current_results))
        
        # Get items for current page
        page_items = self.phpipam_current_results[start_idx:end_idx]
        
        # Display items
        for item in page_items:
            # Create card for each item
            item_card = ctk.CTkFrame(
                self.phpipam_results_scroll, 
                corner_radius=6, 
                fg_color=COLORS["bg_card"]
            )
            item_card.pack(fill="x", pady=3, padx=2)
            
            # Determine if this is a subnet or IP address
            is_subnet = "subnet" in item or "mask" in item
            
            if is_subnet:
                # Subnet card layout
                self._create_subnet_card(item_card, item)
            else:
                # IP address card layout
                self._create_ip_card(item_card, item)
    
    def _create_pagination_controls(self):
        """Create pagination controls for phpIPAM results"""
        total_items = len(self.phpipam_current_results)
        total_pages = (total_items + self.phpipam_items_per_page - 1) // self.phpipam_items_per_page
        
        pagination_frame = ctk.CTkFrame(self.phpipam_results_frame, fg_color="transparent")
        pagination_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        # Previous button
        prev_btn = ctk.CTkButton(
            pagination_frame,
            text="← Previous",
            command=self._prev_page,
            width=100,
            height=32,
            state="normal" if self.phpipam_current_page > 0 else "disabled",
            fg_color=COLORS["neutral"],
            hover_color=COLORS["neutral_hover"]
        )
        prev_btn.pack(side="left", padx=(0, 10))
        
        # Page indicator
        start_idx = self.phpipam_current_page * self.phpipam_items_per_page + 1
        end_idx = min(start_idx + self.phpipam_items_per_page - 1, total_items)
        
        self.phpipam_page_label = ctk.CTkLabel(
            pagination_frame,
            text=f"Showing {start_idx}-{end_idx} of {total_items} • Page {self.phpipam_current_page + 1}/{total_pages}",
            font=ctk.CTkFont(size=11)
        )
        self.phpipam_page_label.pack(side="left", padx=10)
        
        # Next button
        next_btn = ctk.CTkButton(
            pagination_frame,
            text="Next →",
            command=self._next_page,
            width=100,
            height=32,
            state="normal" if self.phpipam_current_page < total_pages - 1 else "disabled",
            fg_color=COLORS["neutral"],
            hover_color=COLORS["neutral_hover"]
        )
        next_btn.pack(side="left")
        
        # Jump to page
        jump_label = ctk.CTkLabel(
            pagination_frame,
            text="Jump to:",
            font=ctk.CTkFont(size=11)
        )
        jump_label.pack(side="right", padx=(10, 5))
        
        self.phpipam_page_entry = ctk.CTkEntry(
            pagination_frame,
            width=60,
            height=32,
            placeholder_text="Page"
        )
        self.phpipam_page_entry.pack(side="right", padx=(0, 5))
        
        jump_btn = ctk.CTkButton(
            pagination_frame,
            text="Go",
            command=self._jump_to_page,
            width=50,
            height=32,
            fg_color=COLORS["primary"],
            hover_color=COLORS["primary_hover"]
        )
        jump_btn.pack(side="right")
    
    def _prev_page(self):
        """Go to previous page"""
        if self.phpipam_current_page > 0:
            self.phpipam_current_page -= 1
            # Clear and recreate
            for widget in self.phpipam_results_frame.winfo_children():
                if widget != self.phpipam_results_frame.winfo_children()[0]:  # Keep title
                    widget.destroy()
            
            # Recreate display
            self._display_phpipam_page()
            self._create_pagination_controls()
    
    def _next_page(self):
        """Go to next page"""
        total_pages = (len(self.phpipam_current_results) + self.phpipam_items_per_page - 1) // self.phpipam_items_per_page
        if self.phpipam_current_page < total_pages - 1:
            self.phpipam_current_page += 1
            # Clear and recreate
            for widget in self.phpipam_results_frame.winfo_children():
                if widget != self.phpipam_results_frame.winfo_children()[0]:  # Keep title
                    widget.destroy()
            
            # Recreate display
            self._display_phpipam_page()
            self._create_pagination_controls()
    
    def _jump_to_page(self):
        """Jump to specific page"""
        try:
            page_num = int(self.phpipam_page_entry.get())
            total_pages = (len(self.phpipam_current_results) + self.phpipam_items_per_page - 1) // self.phpipam_items_per_page
            
            if 1 <= page_num <= total_pages:
                self.phpipam_current_page = page_num - 1
                # Clear and recreate
                for widget in self.phpipam_results_frame.winfo_children():
                    if widget != self.phpipam_results_frame.winfo_children()[0]:  # Keep title
                        widget.destroy()
                
                # Recreate display
                self._display_phpipam_page()
                self._create_pagination_controls()
            else:
                messagebox.showwarning("Invalid Page", f"Please enter a page number between 1 and {total_pages}")
        except ValueError:
            messagebox.showwarning("Invalid Input", "Please enter a valid page number")
    
    def _filter_phpipam_results(self, event=None):
        """Filter displayed phpIPAM results based on search text"""
        if not hasattr(self, 'phpipam_all_results'):
            return
        
        filter_text = self.phpipam_filter_entry.get().lower().strip()
        
        if not filter_text:
            # Show all results
            self.phpipam_current_results = self.phpipam_all_results
        else:
            # Filter results
            filtered = []
            for item in self.phpipam_all_results:
                # Search in all string values
                item_text = " ".join(str(v).lower() for v in item.values() if v)
                if filter_text in item_text:
                    filtered.append(item)
            
            self.phpipam_current_results = filtered
        
        # Reset to first page and redisplay
        self.phpipam_current_page = 0
        
        # Clear and recreate (keeping title)
        for widget in self.phpipam_results_frame.winfo_children():
            if widget != self.phpipam_results_frame.winfo_children()[0]:
                widget.destroy()
        
        # Update count label
        info_frame = ctk.CTkFrame(self.phpipam_results_frame, fg_color="transparent")
        info_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        count_text = f"Showing {len(self.phpipam_current_results)} of {len(self.phpipam_all_results)} result(s)"
        if len(self.phpipam_current_results) > self.phpipam_items_per_page:
            count_text += f" • {self.phpipam_items_per_page} per page"
        
        count_label = ctk.CTkLabel(
            info_frame,
            text=count_text,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLORS["text_secondary"]
        )
        count_label.pack(side="left")
        
        # Recreate scroll frame
        self.phpipam_results_scroll = ctk.CTkScrollableFrame(
            self.phpipam_results_frame,
            height=300
        )
        self.phpipam_results_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 10))
        
        # Display page
        self._display_phpipam_page()
        
        # Add pagination if needed
        if len(self.phpipam_current_results) > self.phpipam_items_per_page:
            self._create_pagination_controls()
    
    def _create_ip_card(self, parent, ip_data):
        """Create formatted card for IP address display"""
        # IP address
        ip = ip_data.get("ip", "N/A")
        
        # Header with IP
        header_frame = ctk.CTkFrame(parent, fg_color="transparent")
        header_frame.pack(fill="x", padx=12, pady=(12, 5))
        
        ip_label = ctk.CTkLabel(
            header_frame,
            text=f"💻 {ip}",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        ip_label.pack(side="left")
        
        # Status badge if available
        if "state" in ip_data:
            state = ip_data["state"]
            state_color = COLORS["online"] if state == "1" else COLORS["offline"]
            state_text = "Active" if state == "1" else "Inactive"
            
            state_label = ctk.CTkLabel(
                header_frame,
                text=state_text,
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color=state_color
            )
            state_label.pack(side="right")
        
        # Hostname
        hostname = ip_data.get("hostname", "")
        if hostname:
            host_label = ctk.CTkLabel(
                parent,
                text=f"🏷️ {hostname}",
                font=ctk.CTkFont(size=11, weight="bold"),
                anchor="w"
            )
            host_label.pack(fill="x", padx=12, pady=(0, 5))
        
        # Description
        description = ip_data.get("description", "")
        if description:
            desc_label = ctk.CTkLabel(
                parent,
                text=description,
                font=ctk.CTkFont(size=10),
                anchor="w",
                text_color=COLORS["text_secondary"]
            )
            desc_label.pack(fill="x", padx=12, pady=(0, 5))
        
        # Additional details
        details = []
        
        if "mac" in ip_data and ip_data["mac"]:
            details.append(f"MAC: {ip_data['mac']}")
        
        if "owner" in ip_data and ip_data["owner"]:
            details.append(f"Owner: {ip_data['owner']}")
        
        if "port" in ip_data and ip_data["port"]:
            details.append(f"Port: {ip_data['port']}")
        
        if details:
            details_text = " • ".join(details)
            details_label = ctk.CTkLabel(
                parent,
                text=details_text,
                font=ctk.CTkFont(size=10),
                anchor="w",
                text_color=COLORS["text_secondary"]
            )
            details_label.pack(fill="x", padx=12, pady=(0, 12))
        else:
            # Add bottom padding
            ctk.CTkFrame(parent, height=5, fg_color="transparent").pack()
    
    def open_live_ping_monitor(self):
        """Open the live ping monitor window"""
        # Now works without matplotlib!
        LivePingMonitorWindow(self)
    
    def change_grid_layout(self, choice):
        """This method is defined in LivePingMonitorWindow class"""
        pass
    
    def toggle_favorite(self, tool_id):
        """Toggle tool as favorite"""
        if tool_id in self.favorite_tools:
            self.favorite_tools.remove(tool_id)
        else:
            self.favorite_tools.add(tool_id)
        
        # Save and update UI
        self.save_window_state()
        self.update_favorites_ui()
        self.update_nav_button_stars()
    
    def update_favorites_ui(self):
        """Update favorites section in sidebar"""
        try:
            # Clear existing buttons
            for widget in self.favorites_buttons_frame.winfo_children():
                widget.destroy()
            
            if self.favorite_tools:
                # Pack frame before first category
                if hasattr(self, 'first_category_label') and self.first_category_label:
                    self.favorites_frame.pack(fill="x", padx=10, pady=(0, 10), before=self.first_category_label)
                else:
                    # Fallback: just pack normally  
                    self.favorites_frame.pack(fill="x", padx=10, pady=(0, 10))
                
                self.favorites_label.pack(anchor="w", pady=(5, 5))
                self.favorites_buttons_frame.pack(fill="x")
                
                # Tool names mapping
                tool_names = {
                    "scanner": "📡 IPv4 Scanner",
                    "portscan": "🔌 Port Scanner",
                    "traceroute": "🛣️ Traceroute",
                    "bandwidth": "⚡ Bandwidth Test",
                    "dns": "🌐 DNS Lookup",
                    "subnet": "🧮 Subnet Calculator",
                    "mac": "💻 MAC Formatter",
                    "compare": "📊 Scan Comparison",
                    "profiles": "📁 Network Profiles",
                    "panos": "🔥 PAN-OS Generator",
                    "phpipam": "📦 phpIPAM",
                }
                
                for tool_id in sorted(self.favorite_tools):
                    btn = StyledButton(
                        self.favorites_buttons_frame,
                        text=tool_names.get(tool_id, tool_id),
                        command=lambda tid=tool_id: self.switch_tool(tid),
                        size="medium",
                        variant="neutral"
                    )
                    btn.pack(fill="x", pady=2)
            else:
                # Completely hide frame when empty (no gap)
                self.favorites_frame.pack_forget()
        except Exception as e:
            print(f"Error updating favorites UI: {e}")
    
    def update_nav_button_stars(self):
        """Update star indicators on navigation buttons"""
        for tool_id, btn in self.nav_buttons.items():
            original_text = btn.cget("text").replace(" ⭐", "")
            if tool_id in self.favorite_tools:
                btn.configure(text=f"{original_text} ⭐")
            else:
                btn.configure(text=original_text)
    
    def show_tool_context_menu(self, event, tool_id):
        """Show context menu for tool (right-click)"""
        import tkinter as tk
        
        context_menu = tk.Menu(self, tearoff=0)
        
        # Favorite/Unfavorite option
        if tool_id in self.favorite_tools:
            context_menu.add_command(
                label="⭐ Remove from Favorites",
                command=lambda: self.toggle_favorite(tool_id)
            )
        else:
            context_menu.add_command(
                label="☆ Add to Favorites",
                command=lambda: self.toggle_favorite(tool_id)
            )
        
        # Show menu at cursor position
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()
    
    def switch_tool(self, tool_id):
        """Switch to a tool"""
        try:
            # Switch to the page
            self.switch_page(tool_id)
        except Exception as e:
            print(f"Error switching tool: {e}")
    
    def on_enter_key(self, event):
        """Handle Enter key press"""
        # Note: Scanner Enter key is now handled within scanner UI
        if self.current_page == "mac":
            if self.format_entries[0].get():
                self.copy_to_clipboard(self.format_entries[0])


class LivePingMonitorWindow(ctk.CTkToplevel):
    """Live Ping Monitor Window with real-time graphs"""
    
    def __init__(self, parent):
        super().__init__(parent)
        
        self.title("Live Ping Monitor")
        self.geometry("1000x600")
        
        # Monitor instance
        self.monitor = LivePingMonitor()
        self.host_widgets = {}  # {ip: {widgets}}
        self.update_interval = 1000  # Update UI every 1 second
        self.updating = False
        
        # Setup UI
        self.setup_ui()
        
        # Bring window to front and focus
        self.attributes('-topmost', True)  # Make window stay on top temporarily
        self.after(100, lambda: self.attributes('-topmost', False))  # Remove topmost after 100ms
        self.lift()
        self.focus_force()
        self.grab_set()  # Make window modal temporarily
        self.after(200, self.grab_release)  # Release after 200ms
    
    def setup_ui(self):
        """Setup the monitor window UI"""
        # Header with controls and legend
        header = ctk.CTkFrame(self, fg_color=COLORS['bg_card'])
        header.pack(fill="x", padx=SPACING['md'], pady=SPACING['md'])
        
        # Control buttons row
        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="left", fill="x", expand=True)
        
        self.stop_btn = StyledButton(
            btn_frame,
            text="⏹ Stop",
            command=self.stop_monitoring,
            size="small",
            variant="danger",
            state="disabled"
        )
        self.stop_btn.pack(side="left", padx=(0, SPACING['sm']))
        
        self.pause_btn = StyledButton(
            btn_frame,
            text="⏸ Pause",
            command=self.pause_monitoring,
            size="small",
            variant="neutral",
            state="disabled"
        )
        self.pause_btn.pack(side="left", padx=(0, SPACING['sm']))
        
        self.resume_btn = StyledButton(
            btn_frame,
            text="▶ Resume",
            command=self.resume_monitoring,
            size="small",
            variant="success",
            state="disabled"
        )
        self.resume_btn.pack(side="left", padx=(0, SPACING['sm']))
        
        self.export_btn = StyledButton(
            btn_frame,
            text="📤 Export",
            command=self.export_data,
            size="small",
            variant="neutral",
            state="disabled"
        )
        self.export_btn.pack(side="left")
        
        # Latency legend on the right
        legend_frame = ctk.CTkFrame(header, fg_color="transparent")
        legend_frame.pack(side="right", padx=SPACING['sm'])
        
        # Green indicator
        green_box = ctk.CTkFrame(legend_frame, fg_color="#00ff00", width=60, height=25, corner_radius=4)
        green_box.pack(side="top", pady=2)
        green_label = ctk.CTkLabel(green_box, text="0-200 ms", font=ctk.CTkFont(size=9), text_color="black")
        green_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Yellow indicator
        yellow_box = ctk.CTkFrame(legend_frame, fg_color="#ffff00", width=60, height=25, corner_radius=4)
        yellow_box.pack(side="top", pady=2)
        yellow_label = ctk.CTkLabel(yellow_box, text="201-500 ms", font=ctk.CTkFont(size=9), text_color="black")
        yellow_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Red indicator
        red_box = ctk.CTkFrame(legend_frame, fg_color="#ff0000", width=60, height=25, corner_radius=4)
        red_box.pack(side="top", pady=2)
        red_label = ctk.CTkLabel(red_box, text="501+ ms", font=ctk.CTkFont(size=9), text_color="white")
        red_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Input section
        input_frame = ctk.CTkFrame(self, fg_color=COLORS['bg_card'])
        input_frame.pack(fill="x", padx=SPACING['md'], pady=(0, SPACING['md']))
        
        self.hosts_entry = StyledEntry(
            input_frame,
            placeholder_text="IPs, CIDRs, Ranges: e.g., 192.168.1.0/24, 10.0.0.1-10.0.0.50, 8.8.8.8"
        )
        self.hosts_entry.pack(side="left", fill="x", expand=True, padx=SPACING['sm'], pady=SPACING['sm'])
        
        self.start_btn = StyledButton(
            input_frame,
            text="▶ Start",
            command=self.start_monitoring,
            size="small",
            variant="success"
        )
        self.start_btn.pack(side="right", padx=SPACING['sm'], pady=SPACING['sm'])
        
        # Table header (reduced height)
        table_header = ctk.CTkFrame(self, fg_color=("gray85", "gray25"), height=28)
        table_header.pack(fill="x", padx=SPACING['md'], pady=(0, 2))
        table_header.pack_propagate(False)
        
        # Column headers
        header_bar = ctk.CTkLabel(table_header, text="", width=8, fg_color="transparent")
        header_bar.pack(side="left", padx=(3, 0))
        
        header_ip = ctk.CTkLabel(table_header, text="IP Address", font=ctk.CTkFont(size=10, weight="bold"), width=140, anchor="w")
        header_ip.pack(side="left", padx=4)
        
        header_hostname = ctk.CTkLabel(table_header, text="Hostname", font=ctk.CTkFont(size=10, weight="bold"), width=180, anchor="w")
        header_hostname.pack(side="left", padx=4)
        
        header_avg = ctk.CTkLabel(table_header, text="Avg", font=ctk.CTkFont(size=10, weight="bold"), width=50, anchor="center")
        header_avg.pack(side="left", padx=4)
        
        header_min = ctk.CTkLabel(table_header, text="Min", font=ctk.CTkFont(size=10, weight="bold"), width=50, anchor="center")
        header_min.pack(side="left", padx=4)
        
        header_cur = ctk.CTkLabel(table_header, text="Current", font=ctk.CTkFont(size=10, weight="bold"), width=60, anchor="center")
        header_cur.pack(side="left", padx=4)
        
        header_graph = ctk.CTkLabel(table_header, text="Graph", font=ctk.CTkFont(size=10, weight="bold"), anchor="center")
        header_graph.pack(side="left", fill="x", expand=True, padx=4)
        
        # Scrollable content area for host rows
        self.scroll_frame = ctk.CTkScrollableFrame(
            self,
            fg_color="transparent"
        )
        self.scroll_frame.pack(fill="both", expand=True, padx=SPACING['md'], pady=(0, SPACING['md']))
        
        # Handle window close
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def parse_host_input(self, input_text):
        """Parse various input formats: IPs, CIDRs, ranges, hostnames"""
        hosts = []
        
        # Split by comma or space
        parts = re.split(r'[,\s]+', input_text)
        
        for part in parts:
            part = part.strip()
            if not part:
                continue
            
            try:
                # Check if it's a CIDR notation (e.g., 192.168.1.0/24)
                if '/' in part:
                    try:
                        network = ipaddress.ip_network(part, strict=False)
                        # Limit to reasonable size to avoid memory issues
                        if network.num_addresses > 1000:
                            messagebox.showwarning(
                                "Too Many Hosts",
                                f"CIDR {part} contains {network.num_addresses} hosts.\n"
                                f"Maximum 1000 hosts allowed per input.\n"
                                f"Please use a smaller subnet."
                            )
                            continue
                        # Add all hosts in the network
                        for ip in network.hosts():
                            hosts.append(str(ip))
                        continue
                    except ValueError:
                        pass
                
                # Check if it's a range (e.g., 192.168.1.1-192.168.1.50)
                if '-' in part:
                    try:
                        start_ip, end_ip = part.split('-', 1)
                        start_ip = start_ip.strip()
                        end_ip = end_ip.strip()
                        
                        # Handle partial end IP (e.g., 192.168.1.1-50)
                        if '.' not in end_ip:
                            # Extract prefix from start IP
                            start_parts = start_ip.split('.')
                            if len(start_parts) == 4:
                                end_ip = '.'.join(start_parts[:3]) + '.' + end_ip
                        
                        # Convert to IP objects
                        start = ipaddress.ip_address(start_ip)
                        end = ipaddress.ip_address(end_ip)
                        
                        # Calculate range size
                        range_size = int(end) - int(start) + 1
                        
                        if range_size > 1000:
                            messagebox.showwarning(
                                "Too Many Hosts",
                                f"Range {part} contains {range_size} hosts.\n"
                                f"Maximum 1000 hosts allowed per input.\n"
                                f"Please use a smaller range."
                            )
                            continue
                        
                        if range_size < 1:
                            messagebox.showwarning(
                                "Invalid Range",
                                f"Invalid range {part}: start IP must be less than or equal to end IP"
                            )
                            continue
                        
                        # Add all IPs in range
                        current = start
                        while current <= end:
                            hosts.append(str(current))
                            current += 1
                        continue
                    except (ValueError, IndexError):
                        pass
                
                # Check if it's a single IP address
                try:
                    ip = ipaddress.ip_address(part)
                    hosts.append(str(ip))
                    continue
                except ValueError:
                    pass
                
                # Assume it's a hostname
                hosts.append(part)
                
            except Exception as e:
                messagebox.showerror("Parse Error", f"Error parsing '{part}':\n{str(e)}")
                continue
        
        return hosts
    
    def start_monitoring(self):
        """Start monitoring the hosts"""
        hosts_input = self.hosts_entry.get().strip()
        
        if not hosts_input:
            messagebox.showwarning("No Hosts", "Please enter at least one IP, CIDR, or range")
            return
        
        # Parse hosts with support for CIDR, ranges, and individual IPs
        hosts = self.parse_host_input(hosts_input)
        
        if not hosts:
            messagebox.showwarning("No Hosts", "No valid hosts found in input")
            return
        
        # Warn if too many hosts
        if len(hosts) > 100:
            result = messagebox.askyesno(
                "Many Hosts",
                f"You are about to monitor {len(hosts)} hosts.\n"
                f"This may impact performance.\n\n"
                f"Continue anyway?"
            )
            if not result:
                return
        
        # Add hosts to monitor
        added_count = 0
        for host in hosts:
            ip = self.monitor.add_host(host)
            if ip:
                self.create_host_widget(ip)
                added_count += 1
        
        if added_count == 0:
            messagebox.showwarning("No Hosts", "Could not add any hosts to monitor")
            return
        
        # Start monitoring
        self.monitor.start_monitoring()
        
        # Update button states
        self.start_btn.configure(state="disabled")
        self.pause_btn.configure(state="normal")
        self.stop_btn.configure(state="normal")
        self.export_btn.configure(state="normal")
        self.hosts_entry.configure(state="disabled")
        
        # Start UI updates
        self.updating = True
        self.update_ui()
    
    def pause_monitoring(self):
        """Pause monitoring"""
        self.monitor.pause_monitoring()
        self.pause_btn.configure(state="disabled")
        self.resume_btn.configure(state="normal")
    
    def resume_monitoring(self):
        """Resume monitoring"""
        self.monitor.resume_monitoring()
        self.pause_btn.configure(state="normal")
        self.resume_btn.configure(state="disabled")
    
    def stop_monitoring(self):
        """Stop monitoring"""
        self.updating = False
        self.monitor.stop_monitoring()
        
        # Update button states
        self.start_btn.configure(state="normal")
        self.pause_btn.configure(state="disabled")
        self.resume_btn.configure(state="disabled")
        self.stop_btn.configure(state="disabled")
        self.hosts_entry.configure(state="normal")
    
    def create_host_widget(self, ip):
        """Create table row widget for a host"""
        host_data = self.monitor.hosts[ip]
        
        # Row container (reduced height for more compact display)
        row = ctk.CTkFrame(self.scroll_frame, fg_color=("gray92", "gray20"), height=38)
        row.pack(fill="x", pady=1)
        row.pack_propagate(False)
        
        # Status bar (colored vertical bar on left)
        status_bar = ctk.CTkFrame(row, fg_color="#808080", width=8)
        status_bar.pack(side="left", fill="y", padx=(3, 0))
        
        # IP Address
        ip_label = ctk.CTkLabel(
            row,
            text=ip,
            font=ctk.CTkFont(size=10),
            width=140,
            anchor="w"
        )
        ip_label.pack(side="left", padx=4)
        
        # Hostname
        hostname_text = host_data.hostname if host_data.hostname else "---"
        hostname_label = ctk.CTkLabel(
            row,
            text=hostname_text,
            font=ctk.CTkFont(size=10),
            width=180,
            anchor="w",
            text_color=COLORS['text_secondary']
        )
        hostname_label.pack(side="left", padx=4)
        
        # Avg latency
        avg_label = ctk.CTkLabel(
            row,
            text="0",
            font=ctk.CTkFont(size=10),
            width=50,
            anchor="center"
        )
        avg_label.pack(side="left", padx=4)
        
        # Min latency
        min_label = ctk.CTkLabel(
            row,
            text="0",
            font=ctk.CTkFont(size=10),
            width=50,
            anchor="center"
        )
        min_label.pack(side="left", padx=4)
        
        # Current latency
        cur_label = ctk.CTkLabel(
            row,
            text="0",
            font=ctk.CTkFont(size=10),
            width=60,
            anchor="center"
        )
        cur_label.pack(side="left", padx=4)
        
        # Graph frame
        graph_frame = ctk.CTkFrame(row, fg_color="transparent")
        graph_frame.pack(side="left", fill="both", expand=True, padx=4)
        
        # Create inline graph using Tkinter Canvas (no matplotlib needed!)
        import tkinter as tk
        canvas = tk.Canvas(
            graph_frame,
            width=260,
            height=28,
            bg='white',
            highlightthickness=0
        )
        canvas.pack(fill="both", expand=True)
        
        # Store references
        self.host_widgets[ip] = {
            'row': row,
            'status_bar': status_bar,
            'ip_label': ip_label,
            'hostname_label': hostname_label,
            'avg_label': avg_label,
            'min_label': min_label,
            'cur_label': cur_label,
            'canvas': canvas,
            'min_value': float('inf'),
            'max_y': 500  # Track max y for scaling
        }
    
    def update_ui(self):
        """Update all host widgets with latest data"""
        if not self.updating:
            return
        
        try:
            all_hosts = self.monitor.get_all_hosts_data()
            
            for ip, host_data in all_hosts.items():
                if ip not in self.host_widgets:
                    continue
                
                widgets = self.host_widgets[ip]
                recent_pings = host_data.get_recent_pings()
                
                # Get current latency (last ping)
                current_latency = 0
                if recent_pings:
                    success, rtt = recent_pings[-1]
                    current_latency = rtt if success else 0
                
                # Calculate stats
                avg_latency = host_data.get_average_latency()
                
                # Track minimum
                if current_latency > 0 and current_latency < widgets['min_value']:
                    widgets['min_value'] = current_latency
                min_latency = widgets['min_value'] if widgets['min_value'] != float('inf') else 0
                
                # Determine status bar color based on current latency
                if current_latency == 0:
                    bar_color = "#ff0000"  # Red - offline
                elif current_latency <= 200:
                    bar_color = "#00ff00"  # Green - good
                elif current_latency <= 500:
                    bar_color = "#ffff00"  # Yellow - moderate
                else:
                    bar_color = "#ff0000"  # Red - high latency
                
                # Update status bar
                widgets['status_bar'].configure(fg_color=bar_color)
                
                # Update statistics labels
                widgets['avg_label'].configure(text=f"{int(avg_latency)}")
                widgets['min_label'].configure(text=f"{int(min_latency)}")
                widgets['cur_label'].configure(text=f"{int(current_latency)}")
                
                # Update graph using Canvas
                canvas = widgets['canvas']
                canvas.delete("all")  # Clear previous drawing
                
                if not recent_pings:
                    continue
                
                # Canvas dimensions
                width = 260
                height = 28
                padding = 2
                
                # Prepare data
                y_data = []
                for success, rtt in recent_pings:
                    if success:
                        y_data.append(rtt)
                    else:
                        y_data.append(None)
                
                # Auto-scale y-axis
                valid_y = [y for y in y_data if y is not None]
                if not valid_y:
                    continue
                
                max_y = max(valid_y)
                widgets['max_y'] = max(500, max_y * 1.1)
                
                # Draw graph
                points = []
                x_step = (width - 2 * padding) / max(len(y_data) - 1, 1)
                
                for i, y in enumerate(y_data):
                    if y is not None:
                        x = padding + i * x_step
                        # Invert y (canvas y=0 is top)
                        y_scaled = height - padding - (y / widgets['max_y']) * (height - 2 * padding)
                        points.append((x, y_scaled))
                
                # Draw line connecting points
                if len(points) > 1:
                    for i in range(len(points) - 1):
                        x1, y1 = points[i]
                        x2, y2 = points[i + 1]
                        canvas.create_line(x1, y1, x2, y2, fill='#0066cc', width=1.5, smooth=True)
                
                # Draw points
                for x, y in points:
                    canvas.create_oval(x-2, y-2, x+2, y+2, fill='#0066cc', outline='#0066cc')
            
            # Schedule next update
            self.after(self.update_interval, self.update_ui)
            
        except Exception as e:
            print(f"Error updating UI: {e}")
            if self.updating:
                self.after(self.update_interval, self.update_ui)
    
    def export_data(self):
        """Export monitoring data to file"""
        if not self.monitor.hosts:
            messagebox.showinfo("No Data", "No monitoring data to export")
            return
        
        # Get export data
        export_text = self.monitor.export_data()
        
        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"ping_monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    f.write(export_text)
                messagebox.showinfo("Export Successful", f"Data exported to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")
    
    def on_closing(self):
        """Handle window close"""
        self.updating = False
        self.monitor.stop_monitoring()
        self.destroy()


if __name__ == "__main__":
    app = NetToolsApp()
    app.mainloop()
