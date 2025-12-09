"""
IPv4 Scanner UI Module
Contains the IPv4 network scanner page implementation with pagination,
import/export, and scan profile management
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import threading
import ipaddress
import csv
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from uuid import uuid4

from design_constants import COLORS, SPACING, RADIUS, FONTS
from ui_components import StyledCard, StyledButton, StyledEntry, ResultRow, SubTitle


class ScannerUI:
    """IPv4 Scanner page UI implementation"""
    
    def __init__(self, app):
        """
        Initialize Scanner UI
        
        Args:
            app: Reference to main NetToolsApp instance
        """
        self.app = app

    def create_content(self, parent):
        """Create IPv4 Scanner page content"""
        # Input section with styled card
        input_card = StyledCard(parent)
        input_card.pack(fill="x", padx=SPACING['lg'], pady=SPACING['lg'])
        
        # CIDR input
        cidr_label = ctk.CTkLabel(
            input_card,
            text="IPv4 / CIDR / Hostname:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        cidr_label.grid(row=0, column=0, padx=SPACING['md'], pady=SPACING['md'], sticky="w")
        
        self.app.cidr_entry = StyledEntry(
            input_card,
            placeholder_text="e.g., 192.168.1.0/24 or server.example.com"
        )
        self.app.cidr_entry.grid(row=0, column=1, padx=(SPACING['md'], SPACING['xs']), pady=SPACING['md'], sticky="ew")
        input_card.grid_columnconfigure(1, weight=1)
        self.app.cidr_entry.bind('<KeyRelease>', self.update_host_count)
        
        # History button for CIDR
        self.app.cidr_history_btn = StyledButton(
            input_card,
            text="⏱",
            size="small",
            variant="neutral",
            command=self.app.show_cidr_history
        )
        self.app.cidr_history_btn.grid(row=0, column=2, padx=(0, SPACING['md']), pady=SPACING['md'])
        
        self.app.host_count_label = SubTitle(
            input_card,
            text=""
        )
        self.app.host_count_label.grid(row=0, column=3, padx=SPACING['md'], pady=SPACING['md'], sticky="w")
        
        # Aggression selector
        aggro_label = ctk.CTkLabel(input_card, text="Aggressiveness:", font=ctk.CTkFont(size=FONTS['body']))
        aggro_label.grid(row=1, column=0, padx=SPACING['md'], pady=SPACING['md'], sticky="w")
        
        self.app.aggro_selector = ctk.CTkOptionMenu(
            input_card,
            values=["Gentle (longer timeout)", "Medium", "Aggressive (short timeout)"],
            width=250
        )
        self.app.aggro_selector.set("Medium")
        self.app.aggro_selector.grid(row=1, column=1, padx=SPACING['md'], pady=SPACING['md'], sticky="ew")
        
        # Scan buttons (moved to separate row for better layout)
        button_frame = ctk.CTkFrame(input_card, fg_color="transparent")
        button_frame.grid(row=2, column=0, columnspan=4, padx=SPACING['md'], pady=SPACING['md'], sticky="ew")
        
        self.app.start_scan_btn = StyledButton(
            button_frame,
            text="▶ Start Scan",
            command=self.start_scan,
            size="medium",
            variant="primary"
        )
        self.app.start_scan_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.import_list_btn = StyledButton(
            button_frame,
            text="📋 Import IP List",
            command=self.import_ip_list,
            size="medium",
            variant="neutral"
        )
        self.app.import_list_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.live_monitor_btn = StyledButton(
            button_frame,
            text="📊 Live Monitor",
            command=self.app.open_live_ping_monitor,
            size="medium",
            variant="success"
        )
        self.app.live_monitor_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.cancel_scan_btn = StyledButton(
            button_frame,
            text="⏹ Cancel",
            command=self.cancel_scan,
            size="medium",
            variant="danger",
            state="disabled"
        )
        self.app.cancel_scan_btn.pack(side="left", padx=SPACING['xs'])
        
        # Spacer
        ctk.CTkLabel(button_frame, text=" ", width=20).pack(side="left")
        
        # Profile buttons
        self.app.save_profile_btn = StyledButton(
            button_frame,
            text="💾 Save Profile",
            command=self.save_scan_profile_dialog,
            size="small",
            variant="neutral"
        )
        self.app.save_profile_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.load_profile_btn = StyledButton(
            button_frame,
            text="📂 Load Profile",
            command=self.load_scan_profile_dialog,
            size="small",
            variant="neutral"
        )
        self.app.load_profile_btn.pack(side="left", padx=SPACING['xs'])
        
        # Options section
        options_frame = ctk.CTkFrame(parent)
        options_frame.pack(fill="x", padx=15, pady=(0, 15))
        
        self.app.only_responding_check = ctk.CTkCheckBox(
            options_frame,
            text="Show only responding hosts",
            command=self.filter_results
        )
        self.app.only_responding_check.select()  # Check by default
        self.app.only_responding_check.pack(side="left", padx=15, pady=15)
        
        self.app.show_all_btn = StyledButton(
            options_frame,
            text="👁 Show All Addresses",
            command=self.show_all_addresses,
            size="large",
            variant="neutral"
        )
        self.app.show_all_btn.pack(side="left", padx=(SPACING['sm'], SPACING['md']), pady=SPACING['md'])
        
        self.app.export_btn = StyledButton(
            options_frame,
            text="📤 Export Results (Ctrl+E)",
            command=self.export_csv,
            size="xlarge",
            variant="success",
            state="disabled"
        )
        self.app.export_btn.pack(side="right", padx=SPACING['md'], pady=SPACING['md'])
        
        self.app.compare_btn = StyledButton(
            options_frame,
            text="📊 Compare Scans",
            command=self.app.show_scan_comparison,
            size="large",
            variant="primary",
            state="disabled"
        )
        self.app.compare_btn.pack(side="right", padx=(0, SPACING['sm']), pady=SPACING['md'])
        
        # Results section with StyledCard
        results_card = StyledCard(parent)
        results_card.pack(fill="both", expand=True, padx=SPACING['lg'], pady=(0, SPACING['lg']))
        
        # Results header with primary color
        header_frame = ctk.CTkFrame(
            results_card,
            height=45,
            corner_radius=RADIUS['medium'],
            fg_color=COLORS['primary']
        )
        header_frame.pack(fill="x", padx=SPACING['xs'], pady=SPACING['xs'])
        header_frame.pack_propagate(False)
        
        headers = [("●", 50), ("IP Address", 180), ("Hostname/FQDN", 250), ("Status", 150), ("RTT (ms)", 100)]
        for text, width in headers:
            label = ctk.CTkLabel(
                header_frame,
                text=text,
                font=ctk.CTkFont(size=FONTS['body'], weight="bold"),
                text_color="white",
                width=width
            )
            label.pack(side="left", padx=SPACING['sm'], pady=SPACING['sm'])
        
        # Pagination controls
        pagination_frame = ctk.CTkFrame(results_card, fg_color="transparent")
        pagination_frame.pack(fill="x", padx=SPACING['lg'], pady=SPACING['xs'])
        
        self.app.pagination_label = ctk.CTkLabel(
            pagination_frame,
            text="No results",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS['text_secondary']
        )
        self.app.pagination_label.pack(side="left")
        
        # Pagination buttons
        pagination_buttons = ctk.CTkFrame(pagination_frame, fg_color="transparent")
        pagination_buttons.pack(side="right")
        
        self.app.first_page_btn = StyledButton(
            pagination_buttons,
            text="⏮ First",
            command=lambda: self.go_to_page(1),
            size="small",
            variant="neutral"
        )
        self.app.first_page_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.prev_page_btn = StyledButton(
            pagination_buttons,
            text="◀ Prev",
            command=self.previous_page,
            size="small",
            variant="neutral"
        )
        self.app.prev_page_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.page_indicator = ctk.CTkLabel(
            pagination_buttons,
            text="Page 1 of 1",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS['text_secondary']
        )
        self.app.page_indicator.pack(side="left", padx=SPACING['md'])
        
        self.app.next_page_btn = StyledButton(
            pagination_buttons,
            text="Next ▶",
            command=self.next_page,
            size="small",
            variant="neutral"
        )
        self.app.next_page_btn.pack(side="left", padx=SPACING['xs'])
        
        self.app.last_page_btn = StyledButton(
            pagination_buttons,
            text="Last ⏭",
            command=lambda: self.go_to_page(self.scan_total_pages),
            size="small",
            variant="neutral"
        )
        self.app.last_page_btn.pack(side="left", padx=SPACING['xs'])
        
        # Scrollable results
        self.app.results_scrollable = ctk.CTkScrollableFrame(results_card)
        self.app.results_scrollable.pack(fill="both", expand=True, padx=SPACING['xs'], pady=(0, SPACING['xs']))
        
        self.app.result_rows = []
        
        # Bind keyboard shortcuts when scanner is active
        parent.bind('<Control-e>', self.export_csv)  # Ctrl+E for export
        self.app.cidr_entry.bind('<Return>', lambda e: self.start_scan() if self.app.start_scan_btn.cget("state") == "normal" else None)
    
    def update_host_count(self, event=None):
        """Update host count label based on CIDR input"""
        cidr = self.app.cidr_entry.get().strip()
        if not cidr:
            self.app.host_count_label.configure(text="")
            return
        
        try:
            network = ipaddress.ip_network(cidr, strict=False)
            if network.prefixlen >= 31:
                count = 2 if network.prefixlen == 31 else 1
            else:
                count = network.num_addresses - 2  # Exclude network and broadcast
            self.app.host_count_label.configure(text=f"Hosts in range: {count:,}")
        except ValueError:
            self.app.host_count_label.configure(text="")
    def start_scan(self):
        """Start network scan"""
        cidr = self.app.cidr_entry.get().strip()
        
        if not cidr:
            messagebox.showinfo(
                "Information",
                "Please enter an IPv4 address, hostname, or CIDR format."
            )
            return
        
        # Check for large scans
        try:
            network = ipaddress.ip_network(cidr, strict=False)
            if network.prefixlen < 20:  # /19 or larger
                host_count = network.num_addresses - 2 if network.prefixlen < 31 else network.num_addresses
                answer = messagebox.askyesno(
                    "Large Scan Warning",
                    f"This range contains {host_count:,} hosts. The scan may take a long time. Continue?"
                )
                if not answer:
                    return
        except ValueError:
            # Not a valid IP/CIDR, try as hostname
            resolved_ip = self.app.scanner.resolve_hostname_to_ip(cidr)
            if resolved_ip:
                # Replace the input with resolved IP and show message
                messagebox.showinfo(
                    "Hostname Resolved",
                    f"Hostname '{cidr}' resolved to {resolved_ip}"
                )
                cidr = resolved_ip  # Use resolved IP for scanning
            else:
                messagebox.showerror(
                    "Invalid Input",
                    f"Could not resolve '{cidr}' as hostname or parse as IP/CIDR."
                )
                return
        
        # Save to history
        self.app.history.add_cidr(cidr)
        
        # Clear previous results
        for row in self.app.result_rows:
            row.destroy()
        self.app.result_rows = []
        
        # Update UI
        self.app.start_scan_btn.configure(state="disabled")
        self.app.cancel_scan_btn.configure(state="normal")
        self.app.export_btn.configure(state="disabled")
        self.app.progress_bar.pack(side="left", padx=15, pady=5)
        self.app.progress_bar.set(0)
        
        # Set callbacks
        self.app.scanner.progress_callback = self.on_scan_progress
        self.app.scanner.complete_callback = self.on_scan_complete
        
        # Start scan in thread
        aggression = self.app.aggro_selector.get()
        self.app.scan_thread = threading.Thread(
            target=self.app.scanner.scan_network,
            args=(cidr, aggression),
            daemon=True
        )
        self.app.scan_thread.start()
        
        self.app.status_label.configure(text="Scan running...")
    def cancel_scan(self):
        """Cancel ongoing scan"""
        self.app.scanner.cancel_scan()
        self.app.status_label.configure(text="Cancelling scan...")
        self.app.cancel_scan_btn.configure(state="disabled")
    def import_ip_list(self):
        """Import and scan IP list from text or file"""
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("Import IP List")
        dialog.geometry("600x500")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - dialog.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Content frame
        content = ctk.CTkFrame(dialog)
        content.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Title
        title = ctk.CTkLabel(
            content,
            text="Import IP Address List",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold")
        )
        title.pack(pady=(0, SPACING['md']))
        
        # Instructions
        instructions = ctk.CTkLabel(
            content,
            text="Enter one per line or load from file\nSupports: IP addresses, CIDR notation, hostnames/FQDNs, comments (#)",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS['text_secondary']
        )
        instructions.pack(pady=(0, SPACING['md']))
        
        # Text area
        text_frame = ctk.CTkFrame(content)
        text_frame.pack(fill="both", expand=True, pady=(0, SPACING['md']))
        
        ip_textbox = ctk.CTkTextbox(
            text_frame,
            font=ctk.CTkFont(size=FONTS['body'], family="Consolas"),
            wrap="none"
        )
        ip_textbox.pack(fill="both", expand=True)
        
        # Placeholder text
        placeholder = """# Example IP list (supports IPs, CIDR, and hostnames):
192.168.1.1
server1.domain.com
192.168.1.10
10.0.0.0/24
workstation.local
gateway.home.lan
172.16.5.100
# This is a comment"""
        ip_textbox.insert("1.0", placeholder)
        
        # Button frame - organized in rows for better visibility
        button_frame = ctk.CTkFrame(content, fg_color="transparent")
        button_frame.pack(fill="x", pady=(SPACING['md'], 0))
        
        # Top row: Load from file
        top_button_row = ctk.CTkFrame(button_frame, fg_color="transparent")
        top_button_row.pack(fill="x", pady=(0, SPACING['xs']))
        
        # Load from file button
        def load_file():
            filepath = filedialog.askopenfilename(
                title="Select IP List File",
                filetypes=[
                    ("Text files", "*.txt"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )
            if filepath:
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        content = f.read()
                    ip_textbox.delete("1.0", "end")
                    ip_textbox.insert("1.0", content)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to load file: {e}")
        
        load_btn = StyledButton(
            top_button_row,
            text="📁 Load from File",
            command=load_file,
            size="medium",
            variant="neutral"
        )
        load_btn.pack(side="left", fill="x", expand=True, padx=(0, SPACING['xs']))
        
        # Bottom row: action buttons
        bottom_button_row = ctk.CTkFrame(button_frame, fg_color="transparent")
        bottom_button_row.pack(fill="x")
        
        # Preview/Validate button
        def preview_list():
            ip_text = ip_textbox.get("1.0", "end")
            
            if not ip_text.strip() or ip_text.strip() == placeholder.strip():
                messagebox.showwarning("Warning", "Please enter IP addresses or load a file")
                return
            
            # Show processing message
            scan_btn.configure(text="⏳ Resolving...", state="disabled")
            preview_btn.configure(state="disabled")
            dialog.update()
            
            # Parse IPs (with hostname resolution)
            ip_list, resolved_info = self.app.scanner.parse_ip_list(ip_text, resolve_hostnames=True)
            
            # Show resolution results
            result_text = "Resolution Results:\n" + "="*50 + "\n\n"
            success_count = 0
            fail_count = 0
            
            for original, resolved, success in resolved_info:
                if success:
                    if original == resolved:
                        result_text += f"✓ {original}\n"
                    else:
                        result_text += f"✓ {original} → {resolved}\n"
                    success_count += 1
                else:
                    result_text += f"✗ {original} - {resolved}\n"
                    fail_count += 1
            
            result_text += "\n" + "="*50 + "\n"
            result_text += f"Total: {success_count} resolved, {fail_count} failed\n"
            result_text += f"Ready to scan {len(ip_list)} IP addresses"
            
            # Show results dialog
            result_dialog = ctk.CTkToplevel(dialog)
            result_dialog.title("Resolution Results")
            result_dialog.geometry("600x400")
            result_dialog.transient(dialog)
            
            result_frame = ctk.CTkFrame(result_dialog)
            result_frame.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
            
            result_display = ctk.CTkTextbox(
                result_frame,
                font=ctk.CTkFont(size=FONTS['small'], family="Consolas"),
                wrap="none"
            )
            result_display.pack(fill="both", expand=True)
            result_display.insert("1.0", result_text)
            result_display.configure(state="disabled")
            
            btn_frame = ctk.CTkFrame(result_frame, fg_color="transparent")
            btn_frame.pack(fill="x", pady=(SPACING['md'], 0))
            
            def proceed_scan():
                result_dialog.destroy()
                if ip_list:
                    dialog.destroy()
                    
                    # Clear previous results
                    self.app.result_rows = []
                    self.app.all_results = []
                    self.app.scan_current_page = 1
                    for widget in self.app.results_scrollable.winfo_children():
                        widget.destroy()
                    
                    # Pre-populate table with all IPs before scanning
                    for ip_addr in ip_list:
                        placeholder_result = {
                            'ip': ip_addr,
                            'hostname': '...',
                            'status': 'Pending',
                            'rtt': '---',
                            'responding': False
                        }
                        self.add_result_row(placeholder_result)
                    
                    # Update UI
                    self.app.start_scan_btn.configure(state="disabled")
                    self.import_list_btn.configure(state="disabled")
                    self.app.cancel_scan_btn.configure(state="normal")
                    self.app.export_btn.configure(state="disabled")
                    self.app.compare_btn.configure(state="disabled")
                    self.app.progress_bar.set(0)
                    self.app.status_label.configure(text=f"Scanning {len(ip_list)} imported addresses...")
                    self.app.cidr_entry.delete(0, 'end')
                    self.app.cidr_entry.insert(0, f"IP List ({len(ip_list)} addresses)")
                    
                    # Store IP list and create mapping for updates
                    self.app.current_scan_list = ip_list
                    self.app.ip_to_row_index = {ip: idx for idx, ip in enumerate(ip_list)}
                    
                    # Set scanner callbacks (CRITICAL!)
                    self.app.scanner.progress_callback = self.on_scan_progress
                    self.app.scanner.complete_callback = self.on_scan_complete
                    
                    # Start scan in background
                    aggression = self.app.aggro_selector.get()  # Use the correct selector
                    
                    threading.Thread(
                        target=self.app.scanner.scan_ip_list,
                        args=(ip_list, aggression),
                        daemon=True
                    ).start()
                else:
                    messagebox.showwarning("Warning", "No valid IP addresses to scan")
            
            StyledButton(
                btn_frame,
                text="✗ Cancel",
                command=result_dialog.destroy,
                size="medium",
                variant="neutral"
            ).pack(side="left")
            
            if ip_list:
                StyledButton(
                    btn_frame,
                    text=f"▶ Scan {len(ip_list)} IPs",
                    command=proceed_scan,
                    size="large",
                    variant="primary"
                ).pack(side="right")
            
            # Re-enable buttons
            scan_btn.configure(text="▶ Scan IP List", state="normal")
            preview_btn.configure(state="normal")
        
        # Define buttons first (referenced in preview_list)
        preview_btn = None
        scan_btn = None
        
        # Cancel button
        cancel_btn = StyledButton(
            bottom_button_row,
            text="✗ Cancel",
            command=dialog.destroy,
            size="medium",
            variant="neutral"
        )
        cancel_btn.pack(side="left", fill="x", expand=True, padx=(0, SPACING['xs']))
        
        # Preview button
        preview_btn = StyledButton(
            bottom_button_row,
            text="🔍 Preview & Resolve",
            command=preview_list,
            size="medium",
            variant="neutral"
        )
        preview_btn.pack(side="left", fill="x", expand=True, padx=(0, SPACING['xs']))
        
        # Scan button
        scan_btn = StyledButton(
            bottom_button_row,
            text="▶ Scan IP List",
            command=preview_list,
            size="large",
            variant="primary"
        )
        scan_btn.pack(side="left", fill="x", expand=True)
    def on_scan_progress(self, completed, total, result):
        """Handle scan progress update with batching for performance"""
        # Add to buffer
        self.app.update_buffer.append((completed, total, result))
        
        # Update immediately if buffer is full OR it's the last result
        if len(self.app.update_buffer) >= self.app.UPDATE_BATCH_SIZE or completed == total:
            self._flush_update_buffer()
        else:
            # Schedule delayed flush if not already scheduled
            if self.app.update_timer is None:
                self.app.update_timer = self.app.after(self.app.UPDATE_INTERVAL_MS, self._flush_update_buffer)
    
    def _flush_update_buffer(self):
        """Flush buffered updates to UI"""
        if self.app.update_timer:
            self.app.after_cancel(self.app.update_timer)
            self.app.update_timer = None
        
        if not self.app.update_buffer:
            return
        
        # Process all buffered updates
        for completed, total, result in self.app.update_buffer:
            self._update_scan_progress(completed, total, result)
        
        self.app.update_buffer.clear()
    
    def _update_scan_progress(self, completed, total, result):
        """Update scan progress in main thread"""
        # Update progress bar
        progress = completed / total if total > 0 else 0
        self.app.progress_bar.set(progress)
        
        # Show current IP being scanned if available
        current_ip = result['ip'] if result else "..."
        status_text = f"Scanning {current_ip}... ({completed} / {total})"
        
        # Add result to all_results list
        self.app.all_results.append(result)
        
        # Check if this is from an imported list
        if hasattr(self, 'current_scan_list') and self.app.current_scan_list and hasattr(self, 'ip_to_row_index'):
            status_text = f"Scanning imported addresses: {current_ip} ({completed}/{total})"
            self.app.status_label.configure(text=status_text)
            
            # Update existing row instead of adding new one
            ip_addr = result.get('ip', '').strip()
            if ip_addr in self.app.ip_to_row_index:
                row_index = self.app.ip_to_row_index[ip_addr]
                if row_index < len(self.app.result_rows):
                    self.update_result_row(row_index, result)
                    # Update pagination UI
                    self.update_pagination_ui()
                    return  # Don't add a new row
            else:
                # IP not in mapping - this shouldn't happen, but handle it
                print(f"Warning: IP {ip_addr} not found in mapping. Available IPs: {list(self.app.ip_to_row_index.keys())[:5]}")
        else:
            self.app.status_label.configure(text=status_text)
        
        # Only add row if on current page and within page limit
        current_page_start = (self.app.scan_current_page - 1) * self.app.results_per_page
        current_page_end = self.app.scan_current_page * self.app.results_per_page
        result_index = len(self.app.all_results) - 1
        
        if current_page_start <= result_index < current_page_end:
            # Add result row (for regular scans or if IP not found in mapping)
            self.add_result_row(result)
        
        # Update pagination UI
        self.update_pagination_ui()
    def on_scan_complete(self, results, message):
        """Handle scan completion"""
        self.after(0, self._finalize_scan, results, message)
    
    def _finalize_scan(self, results, message):
        """Finalize scan in main thread"""
        self.app.start_scan_btn.configure(state="normal")
        self.app.cancel_scan_btn.configure(state="disabled")
        
        # Clear imported list flags
        if hasattr(self, 'current_scan_list'):
            self.app.current_scan_list = None
        if hasattr(self, 'ip_to_row_index'):
            self.app.ip_to_row_index = None
        
        if len(results) > 0:
            self.app.export_btn.configure(state="normal")
            self.app.compare_btn.configure(state="normal")
            
            # Save scan for comparison
            cidr = self.app.cidr_entry.get().strip()
            scan_id = self.app.scan_manager.add_scan(cidr, results)
            self.app.status_label.configure(text=f"{message} (Saved as {scan_id})")
        else:
            self.app.status_label.configure(text=message)
        
        # Render first page to ensure results are visible
        self.app.scan_current_page = 1
        self.render_current_page()
    def add_result_row(self, result):
        """Add a result row to the display with alternating colors"""
        # Determine if this should be an alternate row
        is_alternate = len(self.app.result_rows) % 2 == 1
        
        # Use ResultRow component with alternating colors
        row_color = ("gray92", "gray15") if is_alternate else COLORS['bg_card']
        row_frame = ResultRow(self.app.results_scrollable, fg_color=row_color)
        row_frame.pack(fill="x", padx=SPACING['xs'], pady=SPACING['xs'])
        row_frame._original_color = row_color
        
        # Status dot with better colors
        status_color = COLORS["online"] if result['status'] == 'Online' else COLORS["offline"]
        dot_label = ctk.CTkLabel(
            row_frame,
            text="●",
            font=ctk.CTkFont(size=16),
            text_color=status_color,
            width=50
        )
        dot_label.pack(side="left", padx=SPACING['sm'])
        
        # IP Address with better font
        ip_label = ctk.CTkLabel(
            row_frame,
            text=result['ip'],
            width=180,
            anchor="w",
            font=ctk.CTkFont(size=FONTS['body'])
        )
        ip_label.pack(side="left", padx=SPACING['sm'])
        
        # Hostname/FQDN column (NEW)
        hostname = result.get('hostname', '')
        hostname_label = ctk.CTkLabel(
            row_frame,
            text=hostname if hostname else "-",
            width=250,
            anchor="w",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS["text_secondary"] if hostname else ("gray70", "gray40")
        )
        hostname_label.pack(side="left", padx=SPACING['sm'])
        
        # Status with color and bold font
        status_label = ctk.CTkLabel(
            row_frame,
            text=result['status'],
            width=150,
            anchor="w",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold"),
            text_color=status_color
        )
        status_label.pack(side="left", padx=SPACING['sm'])
        
        # RTT with subtle color
        rtt_text = result.get('rtt', result.get('response_time', '---'))
        rtt_label = ctk.CTkLabel(
            row_frame,
            text=rtt_text,
            width=100,
            anchor="w",
            font=ctk.CTkFont(size=FONTS['small']),
            text_color=COLORS["text_secondary"]
        )
        rtt_label.pack(side="left", padx=SPACING['sm'])
        
        # Store reference
        row_frame.result_data = result
        # Store references to labels for updating
        row_frame.dot_label = dot_label
        row_frame.ip_label = ip_label
        row_frame.hostname_label = hostname_label
        row_frame.status_label = status_label
        row_frame.rtt_label = rtt_label
        self.app.result_rows.append(row_frame)
    def update_result_row(self, row_index, result):
        """Update an existing result row with new data"""
        if row_index >= len(self.app.result_rows):
            return
        
        row_frame = self.app.result_rows[row_index]
        
        # Update status dot color
        status_color = COLORS["online"] if result['status'] == 'Online' else COLORS["offline"]
        row_frame.dot_label.configure(text_color=status_color)
        
        # Update IP (shouldn't change but for consistency)
        row_frame.ip_label.configure(text=result['ip'])
        
        # Update hostname
        hostname = result.get('hostname', '')
        row_frame.hostname_label.configure(
            text=hostname if hostname else "-",
            text_color=COLORS["text_secondary"] if hostname else ("gray70", "gray40")
        )
        
        # Update status
        row_frame.status_label.configure(
            text=result['status'],
            text_color=status_color
        )
        
        # Update RTT
        rtt_text = result.get('rtt', '---')
        row_frame.rtt_label.configure(text=rtt_text)
        
        # Update stored data
        row_frame.result_data = result
    def go_to_page(self, page_num):
        """Go to specific page"""
        if page_num < 1 or page_num > self.app.scan_total_pages:
            return
        
        self.app.scan_current_page = page_num
        self.render_current_page()
    
    def next_page(self):
        """Go to next page"""
        if self.app.scan_current_page < self.app.scan_total_pages:
            self.go_to_page(self.app.scan_current_page + 1)
    
    def previous_page(self):
        """Go to previous page"""
        if self.app.scan_current_page > 1:
            self.go_to_page(self.app.scan_current_page - 1)
    def render_current_page(self):
        """Render results for current page with filtering"""
        # Clear existing rows
        for widget in self.app.results_scrollable.winfo_children():
            widget.destroy()
        self.app.result_rows.clear()
        
        # Apply filter if checkbox is checked
        only_responding = self.app.only_responding_check.get()
        if only_responding:
            filtered_results = [r for r in self.app.all_results if r['status'] == 'Online']
        else:
            filtered_results = self.app.all_results
        
        # Calculate slice based on filtered results
        start_idx = (self.app.scan_current_page - 1) * self.app.results_per_page
        end_idx = start_idx + self.app.results_per_page
        page_results = filtered_results[start_idx:end_idx]
        
        # Render page results
        for result in page_results:
            self.add_result_row(result)
        
        # Update pagination UI with filtered count
        self.update_pagination_ui(filtered_results)
    def update_pagination_ui(self, filtered_results=None):
        """Update pagination controls"""
        # Use filtered results if provided, otherwise use all results
        if filtered_results is None:
            filtered_results = self.app.all_results
        
        total_results = len(filtered_results)
        self.app.scan_total_pages = max(1, (total_results + self.app.results_per_page - 1) // self.app.results_per_page)
        
        # Ensure current page is within bounds
        if self.app.scan_current_page > self.app.scan_total_pages:
            self.app.scan_current_page = self.app.scan_total_pages
        
        # Update labels
        start_idx = (self.app.scan_current_page - 1) * self.app.results_per_page + 1
        end_idx = min(self.app.scan_current_page * self.app.results_per_page, total_results)
        
        if total_results == 0:
            self.app.pagination_label.configure(text="No results")
            self.app.page_indicator.configure(text="Page 0 of 0")
        else:
            self.app.pagination_label.configure(text=f"Showing {start_idx}-{end_idx} of {total_results} results")
            self.app.page_indicator.configure(text=f"Page {self.app.scan_current_page} of {self.app.scan_total_pages}")
        
        # Enable/disable buttons
        self.app.first_page_btn.configure(state="normal" if self.app.scan_current_page > 1 else "disabled")
        self.app.prev_page_btn.configure(state="normal" if self.app.scan_current_page > 1 else "disabled")
        self.app.next_page_btn.configure(state="normal" if self.app.scan_current_page < self.app.scan_total_pages else "disabled")
        self.app.last_page_btn.configure(state="normal" if self.app.scan_current_page < self.app.scan_total_pages else "disabled")
    def filter_results(self, event=None):
        """Filter displayed results and re-render with pagination"""
        # Re-render current page to apply filter
        self.render_current_page()
    def show_all_addresses(self):
        """Show all addresses (uncheck the filter)"""
        self.app.only_responding_check.deselect()
        self.filter_results()
    def export_csv(self, event=None):
        """Export scanner results in multiple formats with options"""
        if not self.app.all_results:
            messagebox.showinfo("Information", "No data to export.")
            return
        
        # Show export options dialog
        self.show_export_dialog()
    
    def show_export_dialog(self):
        """Show export options dialog"""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Export Options")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 500) // 2
        y = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Content
        content = ctk.CTkFrame(dialog)
        content.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Title
        title = ctk.CTkLabel(
            content,
            text="📤 Export Scan Results",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold")
        )
        title.pack(pady=(0, SPACING['lg']))
        
        # Export scope
        scope_frame = ctk.CTkFrame(content, fg_color="transparent")
        scope_frame.pack(fill="x", pady=SPACING['md'])
        
        scope_label = ctk.CTkLabel(
            scope_frame,
            text="Export Scope:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        scope_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        scope_var = ctk.StringVar(value="all")
        
        all_radio = ctk.CTkRadioButton(
            scope_frame,
            text=f"All Results ({len(self.app.all_results)} total)",
            variable=scope_var,
            value="all"
        )
        all_radio.pack(anchor="w", pady=SPACING['xs'])
        
        current_page_radio = ctk.CTkRadioButton(
            scope_frame,
            text=f"Current Page Only ({len(self.app.result_rows)} results)",
            variable=scope_var,
            value="page"
        )
        current_page_radio.pack(anchor="w", pady=SPACING['xs'])
        
        online_radio = ctk.CTkRadioButton(
            scope_frame,
            text=f"Online Hosts Only ({sum(1 for r in self.app.all_results if r.get('status') == 'Online')} results)",
            variable=scope_var,
            value="online"
        )
        online_radio.pack(anchor="w", pady=SPACING['xs'])
        
        # Format selection
        format_frame = ctk.CTkFrame(content, fg_color="transparent")
        format_frame.pack(fill="x", pady=SPACING['md'])
        
        format_label = ctk.CTkLabel(
            format_frame,
            text="Export Format:",
            font=ctk.CTkFont(size=FONTS['body'], weight="bold")
        )
        format_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        format_var = ctk.StringVar(value="csv")
        
        formats = [
            ("csv", "CSV - Comma Separated Values"),
            ("json", "JSON - JavaScript Object Notation"),
            ("xlsx", "Excel - Microsoft Excel (requires openpyxl)"),
            ("html", "HTML - Web Page Report"),
            ("txt", "TXT - Plain Text"),
            ("xml", "XML - Extensible Markup Language"),
        ]
        
        for value, label in formats:
            radio = ctk.CTkRadioButton(
                format_frame,
                text=label,
                variable=format_var,
                value=value
            )
            radio.pack(anchor="w", pady=SPACING['xs'])
        
        # Buttons
        button_frame = ctk.CTkFrame(content, fg_color="transparent")
        button_frame.pack(fill="x", pady=SPACING['lg'], side="bottom")
        
        cancel_btn = StyledButton(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            size="medium",
            variant="neutral"
        )
        cancel_btn.pack(side="right", padx=SPACING['xs'])
        
        def do_export():
            scope = scope_var.get()
            format_type = format_var.get()
            dialog.destroy()
            self.perform_export(scope, format_type)
        
        export_btn = StyledButton(
            button_frame,
            text="📤 Export",
            command=do_export,
            size="large",
            variant="primary"
        )
        export_btn.pack(side="right")
    
    def perform_export(self, scope, format_type):
        """Perform the actual export"""
        # Filter results based on scope
        if scope == "all":
            results_to_export = self.app.all_results
        elif scope == "page":
            results_to_export = [row.result_data for row in self.app.result_rows if hasattr(row, 'result_data')]
        elif scope == "online":
            results_to_export = [r for r in self.app.all_results if r.get('status') == 'Online']
        else:
            results_to_export = self.app.all_results
        
        if not results_to_export:
            messagebox.showinfo("Information", "No data to export with selected filter.")
            return
        
        # Get desktop path
        desktop = Path.home() / "Desktop"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"NetToolsScan_{timestamp}"
        
        # File extension mapping
        ext_map = {
            "csv": ".csv",
            "json": ".json",
            "xlsx": ".xlsx",
            "html": ".html",
            "txt": ".txt",
            "xml": ".xml"
        }
        
        # Ask for save location
        filepath = filedialog.asksaveasfilename(
            defaultextension=ext_map.get(format_type, ".csv"),
            filetypes=[(f"{format_type.upper()} files", f"*{ext_map.get(format_type, '.csv')}")],
            initialdir=desktop,
            initialfile=default_filename
        )
        
        if not filepath:
            return
        
        try:
            if format_type == "csv":
                self._export_as_csv(filepath, results_to_export)
            elif format_type == "json":
                self._export_as_json(filepath, results_to_export)
            elif format_type == "xlsx":
                self._export_as_excel(filepath, results_to_export)
            elif format_type == "html":
                self._export_as_html(filepath, results_to_export)
            elif format_type == "txt":
                self._export_as_txt(filepath, results_to_export)
            elif format_type == "xml":
                self._export_as_xml(filepath, results_to_export)
            
            messagebox.showinfo(
                "Export Successful",
                f"{len(results_to_export)} results exported to:\n{filepath}"
            )
        except Exception as e:
            messagebox.showerror(
                "Export Error",
                f"Error exporting scan results: {str(e)}"
            )
    
    def _export_as_csv(self, filepath, results):
        """Export results as CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['IP Address', 'Hostname', 'Status', 'Response Time'])
            for result in results:
                writer.writerow([
                    result.get('ip', ''),
                    result.get('hostname', ''),
                    result.get('status', ''),
                    result.get('rtt', '')
                ])
    
    def _export_as_json(self, filepath, results):
        """Export results as JSON"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
    
    def _export_as_excel(self, filepath, results):
        """Export results as Excel (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scan Results"
            
            # Header
            headers = ['IP Address', 'Hostname', 'Status', 'Response Time']
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for result in results:
                ws.append([
                    result.get('ip', ''),
                    result.get('hostname', ''),
                    result.get('status', ''),
                    result.get('rtt', '')
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2
            
            wb.save(filepath)
        except ImportError:
            messagebox.showerror(
                "Module Not Found",
                "Excel export requires 'openpyxl' module.\n\nInstall with: pip install openpyxl"
            )
    
    def _export_as_html(self, filepath, results):
        """Export results as HTML report"""
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>NetTools Scan Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        h1 {{ color: #333; }}
        .info {{ background: #e3f2fd; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; background: white; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        th {{ background: #4472C4; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        tr:hover {{ background: #f5f5f5; }}
        .online {{ color: green; font-weight: bold; }}
        .offline {{ color: red; }}
    </style>
</head>
<body>
    <h1>🔍 NetTools Scan Report</h1>
    <div class="info">
        <strong>Generated:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
        <strong>Total Results:</strong> {len(results)}<br>
        <strong>Online Hosts:</strong> {sum(1 for r in results if r.get('status') == 'Online')}<br>
        <strong>Offline Hosts:</strong> {sum(1 for r in results if r.get('status') == 'Offline')}
    </div>
    <table>
        <thead>
            <tr>
                <th>IP Address</th>
                <th>Hostname</th>
                <th>Status</th>
                <th>Response Time</th>
            </tr>
        </thead>
        <tbody>
"""
        for result in results:
            status_class = "online" if result.get('status') == 'Online' else "offline"
            html_content += f"""
            <tr>
                <td>{result.get('ip', '')}</td>
                <td>{result.get('hostname', '-')}</td>
                <td class="{status_class}">{result.get('status', '')}</td>
                <td>{result.get('rtt', '')}</td>
            </tr>
"""
        
        html_content += """
        </tbody>
    </table>
</body>
</html>
"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _export_as_txt(self, filepath, results):
        """Export results as plain text"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("NetTools Scan Report\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Results: {len(results)}\n\n")
            f.write("-" * 80 + "\n")
            f.write(f"{'IP Address':<20} {'Hostname':<30} {'Status':<10} {'RTT':<10}\n")
            f.write("-" * 80 + "\n")
            for result in results:
                f.write(f"{result.get('ip', ''):<20} {result.get('hostname', '-'):<30} {result.get('status', ''):<10} {result.get('rtt', ''):<10}\n")
    
    def _export_as_xml(self, filepath, results):
        """Export results as XML"""
        xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml_content += '<scan_results>\n'
        xml_content += f'  <metadata>\n'
        xml_content += f'    <timestamp>{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</timestamp>\n'
        xml_content += f'    <total_results>{len(results)}</total_results>\n'
        xml_content += f'  </metadata>\n'
        xml_content += '  <hosts>\n'
        for result in results:
            xml_content += '    <host>\n'
            xml_content += f'      <ip>{result.get("ip", "")}</ip>\n'
            xml_content += f'      <hostname>{result.get("hostname", "")}</hostname>\n'
            xml_content += f'      <status>{result.get("status", "")}</status>\n'
            xml_content += f'      <response_time>{result.get("rtt", "")}</response_time>\n'
            xml_content += '    </host>\n'
        xml_content += '  </hosts>\n'
        xml_content += '</scan_results>\n'
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(xml_content)
    def _export_scan_csv(self, filepath):
        """Export IP scan to CSV format"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['ip', 'hostname', 'status', 'rtt'])
            writer.writeheader()
            writer.writerows(self.app.scanner.results)
    
    def _export_scan_json(self, filepath):
        """Export IP scan to JSON format"""
        export_data = {
            'scan_info': {
                'cidr': self.app.cidr_entry.get(),
                'timestamp': datetime.now().isoformat(),
                'total_hosts': len(self.app.scanner.results),
                'online': sum(1 for r in self.app.scanner.results if r['status'] == 'Online'),
                'offline': sum(1 for r in self.app.scanner.results if r['status'] == 'Offline')
            },
            'results': self.app.scanner.results
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2)
    
    def _export_scan_xml(self, filepath):
        """Export IP scan to XML format"""
        root = ET.Element('ipscan')
        
        # Add scan info
        info = ET.SubElement(root, 'scan_info')
        ET.SubElement(info, 'cidr').text = self.app.cidr_entry.get()
        ET.SubElement(info, 'timestamp').text = datetime.now().isoformat()
        ET.SubElement(info, 'total_hosts').text = str(len(self.app.scanner.results))
        ET.SubElement(info, 'online').text = str(sum(1 for r in self.app.scanner.results if r['status'] == 'Online'))
        ET.SubElement(info, 'offline').text = str(sum(1 for r in self.app.scanner.results if r['status'] == 'Offline'))
        
        # Add results
        results = ET.SubElement(root, 'results')
        for result in self.app.scanner.results:
            host = ET.SubElement(results, 'host')
            ET.SubElement(host, 'ip').text = result['ip']
            ET.SubElement(host, 'hostname').text = result.get('hostname', '')
            ET.SubElement(host, 'status').text = result['status']
            ET.SubElement(host, 'rtt').text = str(result.get('rtt', ''))
        
        # Write to file
        tree = ET.ElementTree(root)
        ET.indent(tree, space='  ')
        tree.write(filepath, encoding='utf-8', xml_declaration=True)
    
    def _export_scan_txt(self, filepath):
        """Export IP scan to plain text format"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"IP Network Scan Results\n")
            f.write(f"=" * 60 + "\n")
            f.write(f"CIDR: {self.app.cidr_entry.get()}\n")
            f.write(f"Scan Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Hosts: {len(self.app.scanner.results)}\n")
            f.write(f"Online: {sum(1 for r in self.app.scanner.results if r['status'] == 'Online')}\n")
            f.write(f"Offline: {sum(1 for r in self.app.scanner.results if r['status'] == 'Offline')}\n")
            f.write(f"=" * 60 + "\n\n")
            
            f.write(f"{'IP Address':<18} {'Hostname':<40} {'Status':<10} {'RTT (ms)':<15}\n")
            f.write(f"{'-' * 18} {'-' * 40} {'-' * 10} {'-' * 15}\n")
            
            for result in self.app.scanner.results:
                rtt_str = str(result.get('rtt', 'N/A'))
                hostname_str = result.get('hostname', '-')
                f.write(f"{result['ip']:<18} {hostname_str:<40} {result['status']:<10} {rtt_str:<15}\n")
    def save_scan_profile_dialog(self):
        """Show dialog to save current scan configuration as a profile"""
        # Get current settings
        current_cidr = self.app.cidr_entry.get()
        current_aggression = self.app.aggro_selector.get()
        
        if not current_cidr:
            messagebox.showinfo("No Configuration", "Please enter a CIDR/IP before saving a profile.")
            return
        
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("Save Scan Profile")
        dialog.geometry("400x250")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 400) // 2
        y = self.winfo_y() + (self.winfo_height() - 250) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Content
        content = ctk.CTkFrame(dialog)
        content.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Title
        title = ctk.CTkLabel(
            content,
            text="💾 Save Scan Profile",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold")
        )
        title.pack(pady=(0, SPACING['lg']))
        
        # Profile name
        name_label = ctk.CTkLabel(content, text="Profile Name:", font=ctk.CTkFont(size=FONTS['body']))
        name_label.pack(anchor="w", pady=(0, SPACING['xs']))
        
        name_entry = StyledEntry(content, placeholder_text="e.g., Home Network, DMZ, Guest VLAN")
        name_entry.pack(fill="x", pady=(0, SPACING['md']))
        name_entry.focus()
        
        # Show current config
        config_frame = ctk.CTkFrame(content, fg_color=("gray85", "gray25"))
        config_frame.pack(fill="x", pady=SPACING['md'])
        
        config_text = f"CIDR: {current_cidr}\nAggression: {current_aggression}"
        config_label = ctk.CTkLabel(
            config_frame,
            text=config_text,
            font=ctk.CTkFont(size=FONTS['small']),
            justify="left"
        )
        config_label.pack(padx=SPACING['md'], pady=SPACING['md'])
        
        # Buttons
        button_frame = ctk.CTkFrame(content, fg_color="transparent")
        button_frame.pack(fill="x", pady=(SPACING['lg'], 0))
        
        def save_profile():
            profile_name = name_entry.get().strip()
            if not profile_name:
                messagebox.showwarning("No Name", "Please enter a profile name.")
                return
            
            # Save profile
            self.app.scan_profiles[profile_name] = {
                'cidr': current_cidr,
                'aggression': current_aggression
            }
            self.app.save_scan_profiles()
            
            messagebox.showinfo("Profile Saved", f"Profile '{profile_name}' saved successfully!")
            dialog.destroy()
        
        cancel_btn = StyledButton(
            button_frame,
            text="Cancel",
            command=dialog.destroy,
            size="medium",
            variant="neutral"
        )
        cancel_btn.pack(side="right", padx=SPACING['xs'])
        
        save_btn = StyledButton(
            button_frame,
            text="💾 Save",
            command=save_profile,
            size="medium",
            variant="primary"
        )
        save_btn.pack(side="right")
        
        # Bind Enter key
        name_entry.bind('<Return>', lambda e: save_profile())
    def load_scan_profile_dialog(self):
        """Show dialog to load a saved scan profile"""
        if not self.app.scan_profiles:
            messagebox.showinfo("No Profiles", "No saved profiles found. Save a profile first!")
            return
        
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title("Load Scan Profile")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 500) // 2
        y = self.winfo_y() + (self.winfo_height() - 400) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Content
        content = ctk.CTkFrame(dialog)
        content.pack(fill="both", expand=True, padx=SPACING['lg'], pady=SPACING['lg'])
        
        # Title
        title = ctk.CTkLabel(
            content,
            text="📂 Load Scan Profile",
            font=ctk.CTkFont(size=FONTS['heading'], weight="bold")
        )
        title.pack(pady=(0, SPACING['lg']))
        
        # Profiles list
        profiles_frame = ctk.CTkScrollableFrame(content)
        profiles_frame.pack(fill="both", expand=True, pady=SPACING['md'])
        
        def load_profile(profile_name):
            profile = self.app.scan_profiles[profile_name]
            self.app.cidr_entry.delete(0, 'end')
            self.app.cidr_entry.insert(0, profile['cidr'])
            self.app.aggro_selector.set(profile['aggression'])
            messagebox.showinfo("Profile Loaded", f"Profile '{profile_name}' loaded successfully!")
            dialog.destroy()
        
        def delete_profile(profile_name):
            if messagebox.askyesno("Delete Profile", f"Delete profile '{profile_name}'?"):
                del self.app.scan_profiles[profile_name]
                self.app.save_scan_profiles()
                dialog.destroy()
                # Reopen dialog to show updated list
                self.load_scan_profile_dialog()
        
        for profile_name, profile_data in self.app.scan_profiles.items():
            profile_frame = ctk.CTkFrame(profiles_frame, fg_color=("gray85", "gray25"))
            profile_frame.pack(fill="x", pady=SPACING['xs'])
            
            # Profile info
            info_frame = ctk.CTkFrame(profile_frame, fg_color="transparent")
            info_frame.pack(side="left", fill="both", expand=True, padx=SPACING['md'], pady=SPACING['md'])
            
            name_label = ctk.CTkLabel(
                info_frame,
                text=profile_name,
                font=ctk.CTkFont(size=FONTS['body'], weight="bold"),
                anchor="w"
            )
            name_label.pack(anchor="w")
            
            details_label = ctk.CTkLabel(
                info_frame,
                text=f"CIDR: {profile_data['cidr']} | Aggression: {profile_data['aggression']}",
                font=ctk.CTkFont(size=FONTS['small']),
                text_color=COLORS['text_secondary'],
                anchor="w"
            )
            details_label.pack(anchor="w")
            
            # Buttons
